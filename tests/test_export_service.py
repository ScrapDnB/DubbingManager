"""Тесты для export_service.py"""

import pytest
from typing import Dict, List, Any
from unittest.mock import Mock

import services.export_service as export_module
from services.export_service import ExportService
from core.models import ExportConfig, ReplicaMergeConfig


class TestExportService:
    """Тесты для сервиса экспорта"""

    @pytest.fixture
    def sample_project_data(self) -> Dict[str, Any]:
        """Пример данных проекта для тестов"""
        return {
            "project_name": "Test Project",
            "actors": {
                "actor1": {"name": "Actor One", "color": "#FF0000", "roles": []},
                "actor2": {"name": "Actor Two", "color": "#00FF00", "roles": []},
            },
            "global_map": {
                "Character1": "actor1",
                "Character2": "actor2",
            }
        }

    @pytest.fixture
    def sample_lines(self) -> List[Dict[str, Any]]:
        """Пример реплик для тестов"""
        return [
            {
                "id": 1,
                "s": 0.0,
                "e": 2.5,
                "char": "Character1",
                "text": "Hello, world!",
                "s_raw": "0:00:00.00"
            },
            {
                "id": 2,
                "s": 3.0,
                "e": 5.5,
                "char": "Character2",
                "text": "Hi there!",
                "s_raw": "0:00:03.00"
            },
            {
                "id": 3,
                "s": 6.0,
                "e": 8.5,
                "char": "Character1",
                "text": "How are you?",
                "s_raw": "0:00:06.00"
            }
        ]

    @pytest.fixture
    def export_config(self) -> Dict[str, Any]:
        """Конфигурация экспорта по умолчанию"""
        return {
            'layout_type': 'Таблица',
            'col_tc': True,
            'col_char': True,
            'col_actor': True,
            'col_text': True,
            'f_time': 21,
            'f_char': 20,
            'f_actor': 14,
            'f_text': 30,
            'use_color': True,
            'open_auto': True,
            'round_time': False,
            'time_display': 'range',
            'allow_edit': True
        }

    @pytest.fixture
    def merge_config(self) -> Dict[str, Any]:
        """Конфигурация объединения реплик"""
        return {
            'merge': True,
            'merge_gap': 120,
            'p_short': 0.5,
            'p_long': 2.0,
            'fps': 25.0
        }

    def test_process_merge_logic_no_merge(
        self,
        sample_lines: List[Dict[str, Any]],
        merge_config: Dict[str, Any]
    ) -> None:
        """Тест: объединение реплик отключено"""
        merge_config['merge'] = False
        service = ExportService({})
        result = service.process_merge_logic(sample_lines, merge_config)
        
        # Все реплики должны остаться отдельными
        assert len(result) == len(sample_lines)
        for item in result:
            assert 'parts' in item
            assert len(item['parts']) == 1

    def test_process_merge_logic_same_character(
        self,
        sample_lines: List[Dict[str, Any]],
        merge_config: Dict[str, Any]
    ) -> None:
        """Тест: объединение реплик одного персонажа"""
        # Создаём реплики одного персонажа с маленьким промежутком
        lines = [
            {
                "id": 1,
                "s": 0.0,
                "e": 2.0,
                "char": "Character1",
                "text": "First",
                "s_raw": "0:00:00.00"
            },
            {
                "id": 2,
                "s": 2.2,  # Промежуток 0.2 сек < merge_gap
                "e": 4.0,
                "char": "Character1",
                "text": "Second",
                "s_raw": "0:00:02.00"
            }
        ]
        
        service = ExportService({})
        result = service.process_merge_logic(lines, merge_config)
        
        # Реплики должны объединиться
        assert len(result) == 1
        assert len(result[0]['parts']) == 2
        # Разделитель " / " добавляется при diff >= p_short (0.5)
        assert "First" in result[0]['text']
        assert "Second" in result[0]['text']

    def test_process_merge_logic_different_characters(
        self,
        sample_lines: List[Dict[str, Any]],
        merge_config: Dict[str, Any]
    ) -> None:
        """Тест: реплики разных персонажей не объединяются"""
        service = ExportService({})
        result = service.process_merge_logic(sample_lines, merge_config)
        
        # Реплики разных персонажей не должны объединяться
        assert len(result) == 3

    def test_process_merge_logic_large_gap(
        self,
        merge_config: Dict[str, Any]
    ) -> None:
        """Тест: большой промежуток между репликами"""
        lines = [
            {
                "id": 1,
                "s": 0.0,
                "e": 2.0,
                "char": "Character1",
                "text": "First",
                "s_raw": "0:00:00.00"
            },
            {
                "id": 2,
                "s": 10.0,  # Большой промежуток
                "e": 12.0,
                "char": "Character1",
                "text": "Second",
                "s_raw": "0:00:10.00"
            }
        ]
        
        service = ExportService({})
        result = service.process_merge_logic(lines, merge_config)
        
        # Реплики должны остаться раздельными из-за большого промежутка
        assert len(result) == 2

    def test_generate_html_basic(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: генерация HTML"""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Таблица",
            is_editable=False
        )
        
        assert "<html>" in html
        assert "</html>" in html
        assert "Test Project - Серия 1" in html
        assert "Character1" in html
        assert "Hello, world!" in html

    def test_generate_html_scenario_layout(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: генерация HTML с layout 'Сценарий 1'"""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Сценарий 1",
            is_editable=False
        )
        
        assert "line-container" in html
        assert "Character1" in html

    def test_generate_html_scenario2_layout(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: генерация HTML с layout 'Сценарий 2'."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Сценарий 2",
            is_editable=False
        )

        assert "script2-container" in html
        assert "script2-time" in html
        assert "script2-char" in html
        assert "Character1" in html
        assert "Hello, world!" in html

    def test_generate_html_scenario3_layout(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: генерация HTML с layout 'Сценарий 3'."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Сценарий 3",
            is_editable=False
        )

        assert "script3-table" in html
        assert "script3-meta-cell" in html
        assert "script3-text-cell" in html
        assert "Character1" in html
        assert "Hello, world!" in html

    def test_generate_html_scenario_respects_time_options(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: сценарная разметка учитывает округление и режим тайминга."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        export_config.update({
            "round_time": True,
            "time_display": "start",
        })

        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Сценарий 1",
            is_editable=False
        )

        assert "[0:00:00]" in html
        assert "0:00:02" not in html
        assert "0:00:00.00" not in html

    def test_generate_html_scenario_respects_actor_visibility(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: сценарная разметка скрывает актёра по настройке."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        export_config["col_actor"] = False

        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Сценарий 1",
            is_editable=False
        )

        assert "Character1" in html
        assert "Hello, world!" in html
        assert "Actor1" not in html
        assert "class='a'" not in html

    def test_generate_html_with_highlight(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: генерация HTML с подсветкой актёров"""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        
        # Подсветка только actor1
        highlight_ids = ["actor1"]
        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            highlight_ids=highlight_ids,
            layout_type="Таблица",
            is_editable=False
        )
        
        assert "highlighted-block" in html

    def test_generate_html_can_use_solid_colors_and_negative_text(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: HTML поддерживает плотную заливку и белый текст."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        export_config.update({
            "soften_colors": False,
            "highlight_negative_ids_export": ["actor1"],
        })

        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            highlight_ids=["actor1"],
            layout_type="Таблица",
            is_editable=False
        )

        assert "background-color:#FF0000; color:#ffffff" in html
        assert ".t {\n            font-family: monospace;" in html
        assert "color: inherit;" in html

    def test_generate_html_scenario_negative_applies_to_timing(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: негатив в сценарной разметке распространяется на таймкод."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        export_config.update({
            "soften_colors": False,
            "highlight_negative_ids_export": ["actor1"],
        })

        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            highlight_ids=["actor1"],
            layout_type="Сценарий 1",
            is_editable=False
        )

        assert "border-left-color:#FF0000; color:#ffffff" in html
        assert "<span class='t'>" in html
        assert "color: inherit;" in html

    def test_generate_html_escapes_user_content(
        self,
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: HTML экранирует пользовательский текст и атрибуты."""
        project_data = {
            "project_name": "Project <unsafe>",
            "actors": {
                "actor1": {
                    "name": "Actor <b>One</b>",
                    "color": "#FF0000",
                    "roles": []
                },
            },
            "global_map": {
                "Hero<script>": "actor1",
            }
        }
        processed = [{
            "id": "line",
            "s": 0.0,
            "e": 1.0,
            "char": "Hero<script>",
            "text": "<img src=x onerror=alert(1)> & line",
            "s_raw": "0:00<bad>",
            "parts": [{
                "id": "p' onclick='bad",
                "text": "<img src=x onerror=alert(1)> & part",
                "sep": "<sep>"
            }]
        }]

        html = ExportService(project_data).generate_html(
            ep="1<bad>",
            processed=processed,
            cfg=export_config,
            layout_type="Таблица",
            is_editable=True
        )

        assert "Project &lt;unsafe&gt; - Серия 1&lt;bad&gt;" in html
        assert "Hero&lt;script&gt;" in html
        assert "Actor &lt;b&gt;One&lt;/b&gt;" in html
        assert "&lt;img src=x onerror=alert(1)&gt; &amp; part" in html
        assert "id='p&#x27; onclick=&#x27;bad'" in html
        assert "<img src=x" not in html
        assert "Hero<script>" not in html

    def test_generate_html_table_respects_columns_and_rounded_time(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: HTML-таблица учитывает колонки и округление времени."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {"merge": False})
        export_config.update({
            "col_tc": True,
            "col_char": False,
            "col_actor": False,
            "col_text": True,
            "round_time": True,
        })

        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Таблица",
            is_editable=False
        )

        assert "<th>Время</th><th>Текст</th>" in html
        assert "<th>Персонаж</th>" not in html
        assert "<th>Актер</th>" not in html
        assert "0:00:00<span class='time-sep'>-</span>0:00:02" in html
        assert "0:00:00.00" not in html

    def test_generate_html_table_uses_compact_column_layout(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: HTML-таблица отдаёт больше места реплике."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {"merge": False})

        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Таблица",
            is_editable=False
        )

        assert "<colgroup><col class='col-t'><col class='col-c'><col class='col-a'><col class='col-txt'></colgroup>" in html
        assert ".col-t {\n            width: clamp(5.04rem, 7.35vw, 7.00rem);" in html
        assert ".col-c {\n            width: clamp(7.20rem, 10.50vw, 10.00rem);" in html
        assert ".col-a {\n            width: clamp(6.12rem, 8.93vw, 8.50rem);" in html

    def test_generate_html_table_can_show_start_time_only(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: HTML-таблица может показывать только начало реплики."""
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {"merge": False})
        export_config.update({
            "col_tc": True,
            "round_time": True,
            "time_display": "start",
        })

        html = service.generate_html(
            ep="1",
            processed=processed,
            cfg=export_config,
            layout_type="Таблица",
            is_editable=False
        )

        assert "0:00:00</td>" in html
        assert "<span class='time-sep'>-</span>" not in html

    def test_process_merge_logic_keeps_working_text_lines(self) -> None:
        """Тест: рабочие тексты не объединяются повторно"""
        service = ExportService({})
        lines = [
            {"id": 0, "s": 1.0, "e": 2.0, "char": "Hero", "text": "One", "_working_text": True},
            {"id": 1, "s": 2.1, "e": 3.0, "char": "Hero", "text": "Two", "_working_text": True},
        ]

        result = service.process_merge_logic(
            lines,
            {"merge": True, "merge_gap": 100, "fps": 25}
        )

        assert len(result) == 2
        assert result[0]["text"] == "One"
        assert result[1]["text"] == "Two"

    def test_create_excel_book_respects_highlight_filter(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: Excel подсвечивает только выбранных актёров"""
        pytest.importorskip("openpyxl")
        export_config["highlight_ids_export"] = ["actor1"]

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        wb = service.create_excel_book({"1": processed}, export_config)
        ws = wb["серия (1)"]

        actor1_fill = ws.cell(row=2, column=1).fill.start_color.rgb
        actor2_fill = ws.cell(row=3, column=1).fill.start_color.rgb

        assert actor1_fill == "FFFFC7C7"
        assert actor2_fill in ("FFFFFFFF", "00FFFFFF")

    def test_create_excel_book_accepts_non_numeric_episode_ids(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: Excel экспорт не требует числовых id серий."""
        pytest.importorskip("openpyxl")
        service = ExportService(sample_project_data)
        line = {
            "id": 1,
            "s": 0.0,
            "e": 2.0,
            "char": "Character1",
            "text": "Hello",
            "s_raw": "0:00:00.00"
        }

        wb = service.create_excel_book({
            "pilot": [line],
            "1A": [line],
            "S01E02": [line],
        })

        assert "серия (1A)" in wb.sheetnames
        assert "серия (S01E02)" in wb.sheetnames
        assert "серия (pilot)" in wb.sheetnames

    def test_create_excel_book_can_show_start_time_only(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: Excel может показывать только начало реплики."""
        pytest.importorskip("openpyxl")
        export_config.update({
            "round_time": True,
            "time_display": "start",
        })

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        wb = service.create_excel_book({"1": processed}, export_config)
        ws = wb["серия (1)"]

        assert ws.cell(row=2, column=2).value == "0:00:00"
        assert ws.cell(row=2, column=2).value != "0:00:00-0:00:02"

    def test_create_docx_document_respects_table_columns_and_timing(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX экспорт создаёт таблицу с выбранными колонками."""
        pytest.importorskip("docx")
        export_config.update({
            "col_char": False,
            "col_actor": False,
            "round_time": True,
            "time_display": "range",
        })

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)
        table = document.tables[0]

        headers = [cell.text for cell in table.rows[0].cells]
        first_row = [cell.text for cell in table.rows[1].cells]

        assert headers == ["Время", "Реплика"]
        assert first_row[0] == "0:00:00\n0:00:02"
        assert first_row[1] == "Hello, world!"

    def test_create_docx_document_uses_compact_metadata_columns(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX оставляет больше места колонке реплики."""
        pytest.importorskip("docx")
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)
        first_row = document.tables[0].rows[1].cells

        widths = [
            cell._tc.tcPr.tcW.w
            for cell in first_row
        ]

        assert widths == [714, 1020, 867, 7603]

        grid_widths = [
            int(grid_col.get(export_module.qn('w:w')))
            for grid_col in document.tables[0]._tbl.tblGrid.gridCol_lst
        ]
        assert grid_widths == widths

    def test_create_docx_document_uses_portrait_orientation(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX экспорт использует вертикальную ориентацию листа."""
        pytest.importorskip("docx")
        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)
        section = document.sections[0]

        assert section.page_width < section.page_height

    def test_create_docx_document_can_show_start_time_only(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX экспорт умеет показывать только начало реплики."""
        pytest.importorskip("docx")
        export_config.update({
            "round_time": True,
            "time_display": "start",
        })

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)

        assert document.tables[0].rows[1].cells[0].text == "0:00:00"

    def test_create_docx_document_can_use_solid_negative_highlight(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX поддерживает плотный цвет и белый текст."""
        pytest.importorskip("docx")
        export_config.update({
            "soften_colors": False,
            "highlight_ids_export": ["actor1"],
            "highlight_negative_ids_export": ["actor1"],
        })

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)
        first_cell = document.tables[0].rows[1].cells[0]

        shading = first_cell._tc.tcPr.find(export_module.qn('w:shd'))
        assert shading.get(export_module.qn('w:fill')) == "FF0000"
        assert str(first_cell.paragraphs[0].runs[-1].font.color.rgb) == "FFFFFF"

    def test_create_docx_document_can_use_scenario1_layout(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX поддерживает сценарную разметку 1."""
        pytest.importorskip("docx")
        export_config["layout_type"] = "Сценарий 1"

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)
        first_block = document.tables[0]
        first_cell = first_block.rows[0].cells[0]

        assert len(first_block.rows) == 1
        assert "Character1" in first_cell.text
        assert "[0:00:00,000 - 0:00:02,500]" in first_cell.text
        assert "(Actor One)" in first_cell.text
        assert "Hello, world!" in first_cell.text

    def test_create_docx_document_can_use_scenario2_layout(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX поддерживает сценарную разметку 2."""
        pytest.importorskip("docx")
        export_config.update({
            "layout_type": "Сценарий 2",
            "round_time": True,
        })

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)
        first_cell = document.tables[0].rows[0].cells[0]

        assert "0:00:00\n0:00:02" in first_cell.text
        assert "CHARACTER1" in first_cell.text
        assert "Actor One" in first_cell.text
        assert "Hello, world!" in first_cell.text

    def test_create_docx_document_can_use_scenario3_layout(
        self,
        sample_project_data: Dict[str, Any],
        sample_lines: List[Dict[str, Any]],
        export_config: Dict[str, Any]
    ) -> None:
        """Тест: DOCX поддерживает двухколоночную сценарную разметку."""
        pytest.importorskip("docx")
        export_config.update({
            "layout_type": "Сценарий 3",
            "round_time": True,
        })

        service = ExportService(sample_project_data)
        processed = service.process_merge_logic(sample_lines, {'merge': False})
        document = service.create_docx_document({"1": processed}, export_config)
        table = document.tables[0]

        assert [cell.text for cell in table.rows[0].cells] == [
            "Кто / когда", "Реплика"
        ]
        assert "0:00:00\n0:00:02" in table.rows[1].cells[0].text
        assert "CHARACTER1" in table.rows[1].cells[0].text
        assert "Actor One" in table.rows[1].cells[0].text
        assert table.rows[1].cells[1].text == "Hello, world!"

    def test_export_batch_writes_shared_excel_once(
        self,
        sample_project_data: Dict[str, Any],
        monkeypatch,
        tmp_path
    ) -> None:
        """Тест: общий Excel при пакетном экспорте создаётся один раз."""
        pytest.importorskip("openpyxl")
        sample_project_data["export_config"] = {}
        sample_project_data["replica_merge_config"] = {"merge": False}
        service = ExportService(sample_project_data)
        calls = []

        def fake_export_to_excel(
            ep, lines, cfg, save_path, all_episodes=None, merge_cfg=None
        ):
            calls.append((ep, save_path, list((all_episodes or {}).keys())))
            return True, "ok"

        monkeypatch.setattr(service, "export_to_excel", fake_export_to_excel)
        monkeypatch.setattr(export_module.sys, "platform", "darwin")
        monkeypatch.setattr(export_module.os, "system", lambda command: 0)

        def get_lines(ep):
            return [{
                "id": 1,
                "s": 0.0,
                "e": 1.0,
                "char": "Character1",
                "text": f"Line {ep}",
                "s_raw": "0:00:00.00"
            }]

        success, message = service.export_batch(
            episodes={"1": "one.ass", "2": "two.ass", "3": "three.ass"},
            get_lines_callback=get_lines,
            do_html=False,
            do_xls=True,
            folder=str(tmp_path)
        )

        assert success is True
        assert message == "Экспортировано файлов: 1"
        assert len(calls) == 1
        assert calls[0][2] == ["1", "2", "3"]

    def test_export_batch_writes_docx_per_episode(
        self,
        sample_project_data: Dict[str, Any],
        monkeypatch,
        tmp_path
    ) -> None:
        """Тест: DOCX при пакетном экспорте создаётся по файлу на серию."""
        pytest.importorskip("docx")
        sample_project_data["export_config"] = {}
        sample_project_data["replica_merge_config"] = {"merge": False}
        service = ExportService(sample_project_data)
        monkeypatch.setattr(export_module.sys, "platform", "darwin")
        monkeypatch.setattr(export_module.os, "system", lambda command: 0)

        def get_lines(ep):
            return [{
                "id": 1,
                "s": 0.0,
                "e": 1.0,
                "char": "Character1",
                "text": f"Line {ep}",
                "s_raw": "0:00:00.00"
            }]

        success, message = service.export_batch(
            episodes={"1": "one.ass", "2": "two.ass"},
            get_lines_callback=get_lines,
            do_html=False,
            do_docx=True,
            folder=str(tmp_path)
        )

        assert success is True
        assert message == "Экспортировано файлов: 2"
        assert (tmp_path / "Test Project - Ep1.docx").exists()
        assert (tmp_path / "Test Project - Ep2.docx").exists()
        assert not (tmp_path / "Test Project - Все серии.docx").exists()

    def test_export_batch_respects_open_auto(
        self,
        sample_project_data: Dict[str, Any],
        monkeypatch,
        tmp_path
    ) -> None:
        """Тест: пакетный экспорт не открывает папку при open_auto=False."""
        sample_project_data["export_config"] = {"open_auto": False}
        sample_project_data["replica_merge_config"] = {"merge": False}
        service = ExportService(sample_project_data)
        system_open = Mock()
        monkeypatch.setattr(export_module.sys, "platform", "darwin")
        monkeypatch.setattr(export_module.os, "system", system_open)

        success, message = service.export_batch(
            episodes={"1": "one.ass"},
            get_lines_callback=lambda ep: [{
                "id": 1,
                "s": 0.0,
                "e": 1.0,
                "char": "Character1",
                "text": "Line",
                "s_raw": "0:00:00.00"
            }],
            do_html=True,
            folder=str(tmp_path)
        )

        assert success is True
        assert message == "Экспортировано файлов: 1"
        system_open.assert_not_called()

    def test_generate_reaper_rpp_uses_shared_generator_and_merge_config(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: RPP экспорт использует общую генерацию и настройки слияния."""
        service = ExportService(sample_project_data)
        lines = [
            {
                "id": 1,
                "s": 0.0,
                "e": 1.0,
                "char": "Character1",
                "text": "First",
                "s_raw": "0:00:00.00"
            },
            {
                "id": 2,
                "s": 1.1,
                "e": 2.0,
                "char": "Character1",
                "text": "Second",
                "s_raw": "0:00:01.10"
            },
        ]

        rpp = service.generate_reaper_rpp(
            "1",
            lines,
            merge_cfg={"merge": True, "merge_gap": 120, "fps": 25}
        )

        assert rpp.startswith('<REAPER_PROJECT 0.1 "7.0"')
        assert rpp.count("  MARKER ") == 2
        assert "Character1: First  Second" in rpp
        assert 'MARKER 1 0.0000 "Character1: First  Second" 1 16777471' in rpp
        assert 'MARKER 1 2.0000 "" 1 16777471' in rpp
        assert 'NAME "Actor One"' in rpp
        assert "PEAKCOL 16777471" in rpp

    def test_generate_reaper_rpp_supports_video_and_phrase_regions(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: RPP умеет видео и регионы точно по началу/концу фразы."""
        service = ExportService(sample_project_data)
        lines = [{
            "id": 1,
            "s": 1.0,
            "e": 1.2,
            "char": "Character1",
            "text": 'Quote "line"\nnext',
            "s_raw": "0:00:01.00"
        }]

        rpp = service.generate_reaper_rpp(
            "1",
            lines,
            merge_cfg={"merge": False},
            video_path="/tmp/video file.mov",
            use_video=True,
            use_regions=True
        )

        assert 'MARKER 1 1.0000 "Character1: Quote \' line\'  next" 1 16777471' in rpp
        assert 'MARKER 1 1.2000 "" 1 16777471' in rpp
        assert 'NAME "VIDEO"' in rpp
        assert 'FILE "/tmp/video file.mov"' in rpp

    def test_generate_reaper_rpp_can_transliterate_actor_track_names(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: RPP умеет транслитерировать имена дорожек актёров."""
        sample_project_data["actors"]["actor1"]["name"] = "Иван Ёжиков"
        service = ExportService(sample_project_data)
        lines = [{
            "id": 1,
            "s": 0.0,
            "e": 1.0,
            "char": "Character1",
            "text": "Line",
            "s_raw": "0:00:00.00"
        }]

        rpp = service.generate_reaper_rpp(
            "1",
            lines,
            merge_cfg={"merge": False},
            transliterate_actor_names=True
        )

        assert 'NAME "Ivan Yozhikov"' in rpp
        assert 'NAME "Иван Ёжиков"' not in rpp
        assert 'MARKER 1 0.0000 "Character1: Line" 1 16777471' in rpp

    def test_save_reaper_rpp_writes_utf8_bom_for_cyrillic(
        self,
        sample_project_data: Dict[str, Any],
        tmp_path
    ) -> None:
        """Тест: RPP сохраняется с UTF-8 BOM для корректной кириллицы в Reaper."""
        service = ExportService(sample_project_data)
        save_path = tmp_path / "regions.rpp"
        content = '<REAPER_PROJECT 0.1 "7.0"\n  MARKER 1 0.0000 "Герой: Привет" 1 0\n>'

        service.save_reaper_rpp(str(save_path), content)

        raw = save_path.read_bytes()
        assert raw.startswith(b'\xef\xbb\xbf')
        assert save_path.read_text(encoding='utf-8-sig') == content

    def test_get_reaper_rpp_preview_summarizes_export(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: предпросмотр RPP показывает регионы, актёров и ошибки тайминга."""
        service = ExportService(sample_project_data)
        lines = [
            {
                "id": 1,
                "s": 0.0,
                "e": 1.0,
                "char": "Character1",
                "text": "First",
                "s_raw": "0:00:00.00"
            },
            {
                "id": 2,
                "s": 2.0,
                "e": 2.0,
                "char": "Character2",
                "text": "Broken",
                "s_raw": "0:00:02.00"
            },
        ]

        preview = service.get_reaper_rpp_preview(
            "1",
            lines,
            merge_cfg={"merge": False},
            video_path="/tmp/video.mov",
            use_video=True,
            use_regions=True
        )

        assert preview["regions"] == 2
        assert preview["tracks"] == 2
        assert preview["actors"] == ["Actor One", "Actor Two"]
        assert preview["video"] is True
        assert preview["invalid_lines"] == 1
        assert "Character1: First" in preview["sample_regions"][0]

    def test_get_reaper_rpp_preview_can_transliterate_actor_names(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: предпросмотр RPP показывает транслитерированные имена."""
        sample_project_data["actors"]["actor1"]["name"] = "Пётр Петров"
        service = ExportService(sample_project_data)
        lines = [{
            "id": 1,
            "s": 0.0,
            "e": 1.0,
            "char": "Character1",
            "text": "Line",
            "s_raw": "0:00:00.00"
        }]

        preview = service.get_reaper_rpp_preview(
            "1",
            lines,
            merge_cfg={"merge": False},
            transliterate_actor_names=True
        )

        assert preview["actors"] == ["Pyotr Petrov"]

    def test_count_words(self) -> None:
        """Тест: подсчёт количества слов"""
        service = ExportService({})
        
        assert service._count_words("Hello world") == 2
        assert service._count_words("One") == 1
        assert service._count_words("") == 0
        assert service._count_words("  Multiple   spaces  ") == 2
        assert service._count_words("New\nlines") == 2

    def test_get_colors_with_color_enabled(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: получение цветов с включенной опцией"""
        service = ExportService(sample_project_data)
        
        actor = {"name": "Test", "color": "#FF0000"}
        bg_color, border_col = service._get_colors(
            use_color=True,
            is_highlighted=True,
            actor=actor
        )
        
        assert bg_color is not None
        assert border_col == "#FF0000"

    def test_get_colors_with_color_disabled(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: получение цветов с отключенной опцией"""
        service = ExportService(sample_project_data)
        
        actor = {"name": "Test", "color": "#FF0000"}
        bg_color, border_col = service._get_colors(
            use_color=False,
            is_highlighted=True,
            actor=actor
        )
        
        assert bg_color == "#ffffff"
        assert border_col == "#eee"


class TestExportServiceExcel:
    """Тесты для Excel экспорта (если openpyxl доступен)"""

    @pytest.fixture
    def sample_project_data(self) -> Dict[str, Any]:
        """Пример данных проекта для тестов"""
        return {
            "project_name": "Test Project",
            "actors": {
                "actor1": {"name": "Actor One", "color": "#FF0000", "roles": []},
            },
            "global_map": {
                "Character1": "actor1",
            }
        }

    def test_create_excel_book_basic(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: создание Excel книги"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        service = ExportService(sample_project_data)
        episodes_data = {
            "1": [
                {
                    "id": 1,
                    "s": 0.0,
                    "e": 2.0,
                    "char": "Character1",
                    "text": "Hello",
                    "s_raw": "0:00:00.00"
                }
            ]
        }
        
        wb = service.create_excel_book(episodes_data)
        
        assert wb is not None
        assert len(wb.sheetnames) >= 1
        assert 'Сводка' in wb.sheetnames or any('серия' in name for name in wb.sheetnames)

    def test_create_excel_book_uses_readable_table_column_widths(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: Excel-таблица даёт больше ширины тексту реплики."""
        service = ExportService(sample_project_data)
        wb = service.create_excel_book({
            "1": [{
                "id": 1,
                "s": 0.0,
                "e": 2.0,
                "char": "Character1",
                "text": "Hello",
                "s_raw": "0:00:00.00"
            }]
        })

        ws = wb["серия (1)"]

        assert ws.column_dimensions["B"].width == 10.5
        assert ws.column_dimensions["C"].width == 18.0
        assert ws.column_dimensions["D"].width == 18.0
        assert ws.column_dimensions["E"].width == 118.0

    def test_create_excel_book_multiple_episodes(
        self,
        sample_project_data: Dict[str, Any]
    ) -> None:
        """Тест: создание Excel книги с несколькими сериями"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        service = ExportService(sample_project_data)
        episodes_data = {
            "1": [{"id": 1, "s": 0.0, "e": 2.0, "char": "Character1", "text": "Ep1", "s_raw": "0:00:00.00"}],
            "2": [{"id": 1, "s": 0.0, "e": 2.0, "char": "Character1", "text": "Ep2", "s_raw": "0:00:00.00"}],
            "3": [{"id": 1, "s": 0.0, "e": 2.0, "char": "Character1", "text": "Ep3", "s_raw": "0:00:00.00"}],
        }
        
        wb = service.create_excel_book(episodes_data)
        
        # Должны быть листы для каждой серии + сводка
        assert len(wb.sheetnames) >= 3
