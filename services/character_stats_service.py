"""Character statistics helpers."""

from collections import defaultdict
from typing import Any, Callable, Dict, List, Tuple

from services.assignment_service import get_actor_for_character
from services.export_service import ExportService
from utils.helpers import natural_sort_key

PROJECT_CASTING_METRICS = {"rings", "lines", "words"}


class CharacterStatsService:
    """Calculate episode and project character statistics."""

    def __init__(self, data_ref: Dict[str, Any]) -> None:
        self.data_ref = data_ref

    def episode_stats(
        self,
        lines: List[Dict[str, Any]],
        merge_gap: int,
        fps: float
    ) -> List[Dict[str, Any]]:
        """Calculate per-character stats for one episode."""
        char_data: Dict[str, Dict[str, Any]] = defaultdict(
            lambda: {"lines": 0, "raw": []}
        )

        for line in lines:
            char = line.get('char', '')
            if char:
                char_data[char]["lines"] += 1
                char_data[char]["raw"].append(line)

        merge_gap_seconds = merge_gap / fps

        stats = []
        for char, info in char_data.items():
            rings = 1
            words = 0
            char_lines = info["raw"]

            if char_lines:
                words = len(char_lines[0]['text'].split())

                for i in range(1, len(char_lines)):
                    if char_lines[i]['s'] - char_lines[i - 1]['e'] >= merge_gap_seconds:
                        rings += 1
                    words += len(char_lines[i]['text'].split())

            stats.append({
                "name": char,
                "lines": info["lines"],
                "rings": rings,
                "words": words
            })

        return stats

    def project_stats(
        self,
        char_name: str,
        get_episode_lines: Callable[[str], List[Dict[str, Any]]],
    ) -> Dict[str, Any]:
        """Calculate per-character stats across all project episodes."""
        result: Dict[str, Any] = {
            "rings": 0,
            "words": 0,
            "episodes": []
        }
        export_service = ExportService(self.data_ref)

        for ep in sorted(
            self.data_ref.get("episodes", {}).keys(),
            key=natural_sort_key
        ):
            lines = get_episode_lines(str(ep))
            if not lines:
                continue

            processed = export_service.process_merge_logic(
                lines,
                self.data_ref.get("replica_merge_config", {})
            )
            ep_rings = 0
            ep_words = 0

            for line in processed:
                if line.get("char") != char_name:
                    continue
                ep_rings += 1
                ep_words += len(line.get("text", "").split())

            if ep_rings:
                result["episodes"].append({
                    "episode": str(ep),
                    "rings": ep_rings,
                    "words": ep_words
                })
                result["rings"] += ep_rings
                result["words"] += ep_words

        return result

    def project_casting_summary_rows(
        self,
        get_episode_lines: Callable[[str], List[Dict[str, Any]]],
        metric: str = "rings",
    ) -> List[List[Any]]:
        """Return casting summary rows by character, actor and episode metric."""
        metric = metric if metric in PROJECT_CASTING_METRICS else "rings"
        episodes = sorted(
            (str(ep) for ep in self.data_ref.get("episodes", {}).keys()),
            key=natural_sort_key
        )
        header: List[Any] = ["Персонаж", "Актёр", *episodes, "Всего"]
        rows_by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}
        export_service = ExportService(self.data_ref)
        actors = self.data_ref.get("actors", {})
        next_order = 0

        for ep in episodes:
            lines = get_episode_lines(ep)
            if not lines:
                continue

            processed = lines
            if metric == "rings":
                processed = export_service.process_merge_logic(
                    lines,
                    self.data_ref.get("replica_merge_config", {})
                )

            for line in processed:
                char = line.get("char", "")
                if not char:
                    continue

                actor_id = get_actor_for_character(self.data_ref, char, ep)
                actor_name = (
                    actors.get(actor_id, {}).get("name", "")
                    if actor_id
                    else ""
                )
                key = (char, actor_name)
                row = rows_by_key.get(key)
                if row is None:
                    row = {
                        "char": char,
                        "actor": actor_name,
                        "first_ep": ep,
                        "order": next_order,
                        "episodes": defaultdict(int),
                    }
                    rows_by_key[key] = row
                    next_order += 1
                row["episodes"][ep] += self._project_casting_metric_value(
                    line,
                    metric
                )

        rows: List[List[Any]] = [header]
        rows_by_first_episode: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for row in rows_by_key.values():
            rows_by_first_episode[row["first_ep"]].append(row)

        for ep in episodes:
            episode_rows = sorted(
                rows_by_first_episode.get(ep, []),
                key=lambda row: row["order"]
            )
            if not episode_rows:
                continue

            if len(rows) > 1:
                rows.append([f"{ep} серия", "", *["" for _ in episodes], 0])

            for row in episode_rows:
                counts = [row["episodes"].get(ep_num, 0) for ep_num in episodes]
                total = sum(counts)
                rows.append([
                    row["char"],
                    row["actor"],
                    *[count if count else "" for count in counts],
                    total,
                ])

        return rows

    def _project_casting_metric_value(
        self,
        line: Dict[str, Any],
        metric: str,
    ) -> int:
        """Return one exported cell contribution for the selected metric."""
        if metric == "words":
            return len(line.get("text", "").split())
        return 1

    def create_project_casting_xlsx(
        self,
        get_episode_lines: Callable[[str], List[Dict[str, Any]]],
        metric: str = "rings",
    ) -> Any:
        """Return a formatted XLSX workbook for Google Sheets import."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError("openpyxl не установлен") from exc

        rows = self.project_casting_summary_rows(get_episode_lines, metric=metric)
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводка"

        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        bold_font = Font(bold=True)
        wrap_alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="left",
        )

        for row in rows:
            ws.append(row)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment

        max_column = ws.max_column
        for cell in ws[1]:
            cell.fill = yellow_fill
            cell.font = bold_font

        for row_idx in range(2, ws.max_row + 1):
            first_value = ws.cell(row=row_idx, column=1).value
            second_value = ws.cell(row=row_idx, column=2).value
            if isinstance(first_value, str) and first_value.endswith(" серия") and not second_value:
                for col_idx in range(1, max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = yellow_fill
                    cell.font = bold_font

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 30
        for col_idx in range(3, max_column):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8
        ws.column_dimensions[get_column_letter(max_column)].width = 10

        ws.freeze_panes = "C2"
        return wb
