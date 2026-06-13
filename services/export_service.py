"""Service for exporting project data."""

import os
import sys
import logging
import re
import subprocess
from typing import Dict, List, Any, Optional, Set, Tuple, Callable

from services.assignment_service import get_actor_for_character
from services.export_layouts import ExportLayoutMixin
from services.pdf_export_service import PdfExportService
from services.replica_merge_service import ReplicaMergeService
from services.reaper_rpp_service import ReaperRppService
from utils.helpers import (
    format_seconds_to_full_tc,
    format_seconds_to_tc,
    format_timing_range,
)
from utils.i18n import translate_source

logger = logging.getLogger(__name__)

try:
    from docx.oxml.ns import qn
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    qn = None
    logger.warning("python-docx not available - DOCX export disabled")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export disabled")


class ExportService(ExportLayoutMixin):
    """Export Service implementation."""

    def __init__(self, project_data: Dict[str, Any]):
        self.project_data = project_data
        self.pdf_export_service = PdfExportService()
        self.replica_merge_service = ReplicaMergeService()
        self.reaper_rpp_service = ReaperRppService(project_data)

    def _open_path(self, path: str) -> None:
        """Open a local path without going through a shell."""
        if sys.platform == 'win32':
            os.startfile(path)
        elif sys.platform == 'darwin':
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)

    def _get_effective_highlight_filter(
        self,
        cfg: Dict[str, Any]
    ) -> Optional[Set[str]]:
        """Return effective highlight filter."""
        highlight_ids = cfg.get('highlight_ids_export')
        if highlight_ids is None:
            return None

        selected_ids = set(highlight_ids)
        all_actor_ids = set(self.project_data.get("actors", {}).keys())
        if selected_ids == all_actor_ids:
            return None

        return selected_ids

    def _format_export_timing(
        self,
        line: Dict[str, Any],
        cfg: Dict[str, Any]
    ) -> str:
        """Format timing according to export settings."""
        start = line.get('s', 0)
        end = line.get('e', 0)
        start_only = cfg.get('time_display', 'range') == 'start'

        if cfg.get('round_time', False):
            start_tc = format_seconds_to_tc(start)
            if start_only:
                return start_tc
            return f"{start_tc}-{format_seconds_to_tc(end)}"

        if start_only:
            return format_seconds_to_full_tc(start)

        return format_timing_range(start, end)

    def _format_table_timing_text(
        self,
        line: Dict[str, Any],
        cfg: Dict[str, Any]
    ) -> str:
        """Format timing for table-like exports."""
        timing = self._format_export_timing(line, cfg)
        if cfg.get('time_display', 'range') == 'start':
            return timing
        return timing.replace('-', '\n')

    def process_merge_logic(
        self,
        lines: List[Dict[str, Any]],
        cfg: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Apply replica merge rules."""
        return self.replica_merge_service.process(lines, cfg)

    def _episode_sort_key(self, ep_num: Any) -> Tuple[Any, ...]:
        """Return a natural sort key for episode identifiers."""
        parts = re.split(r'(\d+)', str(ep_num))
        return tuple(
            (0, int(part)) if part.isdigit() else (1, part.lower())
            for part in parts
            if part != ''
        )

    # ==========================================================================
    # Excel export
    # ==========================================================================

    def _get_times_font(
        self,
        size: float = 14.0,
        bold: bool = False,
        italic: bool = False,
        color: Optional[str] = None
    ) -> Font:
        """Return times font."""
        return Font(
            name='Times New Roman',
            size=size,
            bold=bold,
            italic=italic,
            charset=204,
            color=color
        )

    def _get_thin_border(self) -> Border:
        """Return thin border."""
        side = Side(style='thin', color='00000000')
        return Border(left=side, right=side, top=side, bottom=side)

    def _count_words(self, text: str) -> int:
        """Count words in text."""
        if not text:
            return 0
        # Split on whitespace and count non-empty tokens
        words = re.findall(r'\S+', text.strip())
        return len(words)

    def _apply_cell_styling(
        self,
        cell,
        font_size: float = 14.0,
        wrap_text: bool = False,
        fill_color: Optional[str] = None,
        text_color: Optional[str] = None,
        border: Optional[Border] = None
    ):
        """Apply cell styling."""
        font_color = None
        if text_color:
            font_color = text_color.replace('#', '')
            if len(font_color) == 6:
                font_color = 'FF' + font_color
        cell.font = self._get_times_font(size=font_size, color=font_color)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=wrap_text)
        if border:
            cell.border = border
        if fill_color:
            # Ensure color is in ARGB format (8 chars)
            color = fill_color.replace('#', '')
            if len(color) == 6:
                color = 'FF' + color  # Add alpha channel
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.fill = fill

    def _create_actors_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        episodes_data: Dict[str, List[Dict[str, Any]]],
        cfg: Dict[str, Any]
    ):
        """Create actors summary sheet."""
        # Remove the default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        ws = wb.create_sheet(title='Сводка')

        actors = self.project_data.get('actors', {})
        effective_filter = self._get_effective_highlight_filter(cfg)
        soften_colors = cfg.get('soften_colors', True)

        # Sort episode numbers for the correct column order
        sorted_ep_keys = sorted(episodes_data.keys(), key=self._episode_sort_key)

        # Collect actor statistics
        actor_stats: Dict[str, Dict[str, Any]] = {}
        for actor_id, actor_data in actors.items():
            actor_stats[actor_id] = {
                'name': actor_data.get('name', ''),
                'color': actor_data.get('color', '#FFFFFF'),
                'roles': [],
                'episode_words': {}
            }

        # Count words by episode using the real episode numbers
        for ep_key in sorted_ep_keys:
            lines = episodes_data[ep_key]
            for line in lines:
                char_name = line.get('char', '')
                actor_id = get_actor_for_character(
                    self.project_data, char_name, ep_key
                )
                if actor_id and actor_id in actor_stats:
                    if char_name not in actor_stats[actor_id]['roles']:
                        actor_stats[actor_id]['roles'].append(char_name)
                    if ep_key not in actor_stats[actor_id]['episode_words']:
                        actor_stats[actor_id]['episode_words'][ep_key] = 0
                    actor_stats[actor_id]['episode_words'][ep_key] += self._count_words(line.get('text', ''))

        # Dynamic headers with real episode numbers
        headers = ['Актёр', 'Персонаж']
        for ep_key in sorted_ep_keys:
            headers.append(f'{ep_key} серия')
        headers.append('Всего слов')
        ws.append(headers)

        # Header styles
        header_font = self._get_times_font(size=14.0)
        header_alignment = Alignment(horizontal='left', vertical='top')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment

        # Column widths
        ws.column_dimensions['A'].width = 21.5
        ws.column_dimensions['B'].width = 55.33
        # Dynamic width for episode columns
        for i in range(1, len(sorted_ep_keys) + 2):
            col_letter = openpyxl.utils.get_column_letter(2 + i)
            ws.column_dimensions[col_letter].width = 11.83

        # First-row height
        ws.row_dimensions[1].height = 20.0

        # Fill data
        row_num = 2
        thin_border = self._get_thin_border()
        for actor_id, stats in actor_stats.items():
            if not stats['roles']:  # Skip actors without roles
                continue

            actor_name = stats['name']
            roles_str = ', '.join(stats['roles'])
            episode_words = stats['episode_words']

            # Actor color
            is_highlighted = (
                effective_filter is None or
                actor_id in effective_filter
            )
            if is_highlighted:
                actor_color = stats['color']
                fill_color = (
                    self._docx_soft_fill_color(actor_color)
                    if soften_colors
                    else actor_color.replace('#', '')
                )
            else:
                fill_color = 'FFFFFF'
            text_color = self._negative_text_color(actor_id, cfg, is_highlighted)

            # Actor-name cell
            name_cell = ws.cell(row=row_num, column=1, value=actor_name)
            self._apply_cell_styling(
                name_cell,
                font_size=14.0,
                fill_color=fill_color,
                text_color=text_color,
                border=thin_border
            )

            # Character cell
            roles_cell = ws.cell(row=row_num, column=2, value=roles_str)
            self._apply_cell_styling(roles_cell, font_size=9.0, border=thin_border)

            # Word-count cells by episode in sorted-key order
            total_words = 0
            for ep_idx, ep_key in enumerate(sorted_ep_keys):
                word_count = episode_words.get(ep_key, 0)
                total_words += word_count
                cell = ws.cell(row=row_num, column=3 + ep_idx, value=word_count)
                self._apply_cell_styling(cell, font_size=14.0, border=thin_border)

            # Total word-count cell
            total_cell = ws.cell(row=row_num, column=3 + len(sorted_ep_keys), value=total_words)
            self._apply_cell_styling(total_cell, font_size=14.0, border=thin_border)

            row_num += 1

        # Do not add a filter; sorting is not needed on the actor page

    def _create_episode_sheet(
        self,
        wb: openpyxl.Workbook,
        ep_num: str,
        processed: List[Dict[str, Any]],
        cfg: Dict[str, Any]
    ):
        """Create one episode sheet."""
        sheet_name = f'серия ({ep_num})'
        ws = wb.create_sheet(title=sheet_name)

        actors = self.project_data.get('actors', {})
        use_color = cfg.get('use_color', True)
        soften_colors = cfg.get('soften_colors', True)
        effective_filter = self._get_effective_highlight_filter(cfg)

        # Choose columns from settings
        col_tc = cfg.get('col_tc', True)
        col_char = cfg.get('col_char', True)
        col_actor = cfg.get('col_actor', True)
        col_text = cfg.get('col_text', True)

        # Headers only for selected columns
        headers = [translate_source('Номер')]
        if col_tc:
            headers.append(translate_source('Таймкод'))
        if col_char:
            headers.append(translate_source('Персонаж'))
        if col_actor:
            headers.append(translate_source('Актёр'))
        if col_text:
            headers.append(translate_source('Реплика'))
        ws.append(headers)

        # Header styles
        header_fonts = {
            'номер': self._get_times_font(size=14.0),
            'тайм': self._get_times_font(size=13.0),
            'персонаж': self._get_times_font(size=12.0),
            'актёр': self._get_times_font(size=14.0),
            'реплика': self._get_times_font(size=14.0)
        }
        header_alignment = Alignment(horizontal='left', vertical='top')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_fonts.get(header, self._get_times_font(size=14.0))
            cell.alignment = header_alignment

        # Column widths
        col_widths = {
            'Номер': 6.5,
            'Таймкод': 10.5,
            'Персонаж': 18.0,
            'Актёр': 18.0,
            'Реплика': 118.0
        }
        for col_idx, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_widths.get(header, 13.0)

        # Thin borders for all cells
        thin_border = self._get_thin_border()

        # Fill data
        for row_idx, line in enumerate(processed, 2):
            char_name = line.get('char', '')
            actor_id = get_actor_for_character(
                self.project_data, char_name, ep_num
            )
            actor = actors.get(actor_id, {}) if actor_id else {}
            actor_name = actor.get('name', '-') if actor else '-'

            # Determine the color
            is_highlighted = (
                effective_filter is None or
                actor_id in effective_filter
            )
            if use_color and actor_id and is_highlighted:
                actor_color = actor.get('color', '#FFFFFF')
                color_hex = (
                    self._docx_soft_fill_color(actor_color)
                    if soften_colors
                    else actor_color.replace('#', '')
                )
            else:
                color_hex = 'FFFFFF'
            text_color = self._negative_text_color(
                actor_id,
                cfg,
                is_highlighted
            )

            timing = self._format_export_timing(line, cfg)

            # Row data only for selected columns
            row_data = [row_idx - 1]  # Always include the row number
            if col_tc:
                row_data.append(timing)
            if col_char:
                row_data.append(char_name)
            if col_actor:
                row_data.append(actor_name)
            if col_text:
                row_data.append(line.get('text', ''))

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)

                # Apply styles
                font_size = 14.0
                # Use wrap_text for timing and replica text
                timing_col = 2 if col_tc else None  # Timing column, second after the number
                wrap_text = (timing_col and col == timing_col) or (col_text and col == len(row_data))

                self._apply_cell_styling(
                    cell,
                    font_size=font_size,
                    wrap_text=wrap_text,
                    fill_color=color_hex,
                    text_color=text_color,
                    border=thin_border
                )

                # Set row height for the replica column
                if col_text and col == len(row_data):
                    # Approximate row height based on text length
                    text = value if isinstance(value, str) else ''
                    lines_count = max(1, text.count('\n') + 1 + len(text) // 80)
                    ws.row_dimensions[row_idx].height = min(120, 20 + lines_count * 15)

        # Add a filter
        last_col_letter = openpyxl.utils.get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col_letter}{len(processed) + 1}'

    def create_excel_book(
        self,
        episodes_data: Dict[str, List[Dict[str, Any]]],
        cfg: Optional[Dict[str, Any]] = None
    ) -> Any:
        """Create an Excel workbook with multiple sheets."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not available")

        if cfg is None:
            cfg = self.project_data.get("export_config", {})

        wb = openpyxl.Workbook()

        # Create the actor summary sheet
        self._create_actors_summary_sheet(wb, episodes_data, cfg)

        # Create episode sheets in sorted order
        for ep_num in sorted(episodes_data.keys(), key=self._episode_sort_key):
            lines = episodes_data[ep_num]
            self._create_episode_sheet(wb, ep_num, lines, cfg)

        return wb

    def export_to_docx(
        self,
        ep: str,
        lines: List[Dict[str, Any]],
        cfg: Dict[str, Any],
        save_path: str,
        all_episodes: Optional[Dict[str, List[Dict[str, Any]]]] = None,
        merge_cfg: Optional[Dict[str, Any]] = None
    ) -> Tuple[bool, str]:
        """Export data to a DOCX table document."""
        if not DOCX_AVAILABLE:
            return False, translate_source("python-docx не установлен")

        try:
            episodes_data = all_episodes if all_episodes else {ep: lines}
            if merge_cfg and not all_episodes:
                episodes_data = {
                    ep: self.process_merge_logic(lines, merge_cfg)
                }
            document = self.create_docx_document(episodes_data, cfg)
            document.save(save_path)
            return True, f"{translate_source('Экспортировано в')} {save_path}"
        except Exception as e:
            logger.error(f"DOCX export error: {e}")
            return False, f"{translate_source('Ошибка экспорта DOCX:')} {e}"

    def export_to_pdf(
        self,
        ep: str,
        lines: List[Dict[str, Any]],
        cfg: Dict[str, Any],
        save_path: str,
        all_episodes: Optional[Dict[str, List[Dict[str, Any]]]] = None,
        merge_cfg: Optional[Dict[str, Any]] = None
    ) -> Tuple[bool, str]:
        """Export data to a PDF document."""
        try:
            episodes_data = all_episodes if all_episodes else {ep: lines}
            if merge_cfg is None:
                merge_cfg = self.project_data.get("replica_merge_config", {})

            html_parts = []
            for ep_num in sorted(episodes_data.keys(), key=self._episode_sort_key):
                ep_lines = episodes_data[ep_num]
                processed = self.process_merge_logic(ep_lines, merge_cfg)
                html_parts.append(
                    self.generate_html(
                        ep_num,
                        processed,
                        cfg,
                        cfg.get('highlight_ids_export'),
                        layout_type=cfg.get('layout_type', 'Таблица'),
                        is_editable=False
                    )
                )

            html = "\n<div style='page-break-after: always'></div>\n".join(
                html_parts
            )
            self.pdf_export_service.render_html_to_pdf(html, save_path)
            return True, f"{translate_source('PDF сохранён:')} {save_path}"
        except Exception as e:
            logger.error(f"PDF export error: {e}")
            return False, f"{translate_source('Ошибка экспорта PDF:')} {e}"

    def export_to_excel(
        self,
        ep: str,
        lines: List[Dict[str, Any]],
        cfg: Dict[str, Any],
        save_path: str,
        all_episodes: Optional[Dict[str, List[Dict[str, Any]]]] = None,
        merge_cfg: Optional[Dict[str, Any]] = None
    ) -> Tuple[bool, str]:
        """Export data to an Excel file."""
        if not EXCEL_AVAILABLE:
            return False, translate_source("openpyxl не установлен")

        try:
            # Use all episodes for the summary when provided
            if all_episodes:
                episodes_data = all_episodes
            else:
                episodes_data = {ep: lines}

            # Use merge_cfg for replica processing and cfg for formatting
            if merge_cfg is None:
                merge_cfg = self.project_data.get("replica_merge_config", {})

            processed_episodes = {}
            for ep_num, ep_lines in episodes_data.items():
                processed_episodes[ep_num] = self.process_merge_logic(ep_lines, merge_cfg)

            wb = self.create_excel_book(processed_episodes, cfg)
            wb.save(save_path)
            return True, f"{translate_source('Excel сохранён:')} {save_path}"
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            return False, f"{translate_source('Ошибка экспорта:')} {e}"

    # ==========================================================================
    # Reaper RPP export
    # ==========================================================================

    def _hex_to_reaper_color(self, hex_color: str) -> int:
        """Hex to reaper color."""
        return self.reaper_rpp_service._hex_to_reaper_color(hex_color)

    def _escape_rpp_text(self, text: Any) -> str:
        """Escape rpp text."""
        return self.reaper_rpp_service._escape_rpp_text(text)

    def _transliterate_cyrillic(self, text: Any) -> str:
        """Transliterate Cyrillic text to Latin for DAW-friendly track names."""
        return self.reaper_rpp_service._transliterate_cyrillic(text)

    def _reaper_actor_name(
        self,
        actor: Dict[str, Any],
        transliterate_actor_names: bool = False
    ) -> str:
        """Return actor name prepared for Reaper track display."""
        return self.reaper_rpp_service._reaper_actor_name(
            actor,
            transliterate_actor_names
        )

    def save_reaper_rpp(self, save_path: str, rpp_content: str) -> None:
        """Save an RPP file with an encoding Reaper reads reliably."""
        self.reaper_rpp_service.save(save_path, rpp_content)

    def generate_reaper_rpp(
        self,
        ep: str,
        lines: List[Dict[str, Any]],
        merge_cfg: Optional[Dict[str, Any]] = None,
        video_path: Optional[str] = None,
        use_video: bool = False,
        use_regions: bool = True,
        transliterate_actor_names: bool = False
    ) -> str:
        """Generate a Reaper RPP project from episode lines."""
        return self.reaper_rpp_service.generate(
            ep,
            lines,
            merge_cfg,
            video_path,
            use_video,
            use_regions,
            transliterate_actor_names
        )

    def get_reaper_rpp_preview(
        self,
        ep: str,
        lines: List[Dict[str, Any]],
        merge_cfg: Optional[Dict[str, Any]] = None,
        video_path: Optional[str] = None,
        use_video: bool = False,
        use_regions: bool = True,
        transliterate_actor_names: bool = False
    ) -> Dict[str, Any]:
        """Return a user-facing preview summary for RPP export."""
        return self.reaper_rpp_service.preview(
            ep,
            lines,
            merge_cfg,
            video_path,
            use_video,
            use_regions,
            transliterate_actor_names
        )

    # ==========================================================================
    # Batch export
    # ==========================================================================

    def export_batch(
        self,
        episodes: Dict[str, str],
        get_lines_callback,
        do_html: bool = True,
        do_xls: bool = False,
        do_docx: bool = False,
        do_pdf: bool = False,
        folder: str = None,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> Tuple[bool, str]:
        """Export several episodes in one batch."""
        if not folder:
            return False, translate_source("Папка для экспорта не указана")

        cfg = self.project_data["export_config"]
        merge_cfg = self.project_data.get("replica_merge_config", {})
        project_name = self.project_data.get('project_name', 'Project')
        exported_count = 0
        total_episodes = len(episodes)

        try:
            # Collect all episodes for the Excel summary
            all_episodes_data = {}
            if do_xls:
                for ep, path in episodes.items():
                    lines = get_lines_callback(ep)
                    if lines:
                        all_episodes_data[ep] = self.process_merge_logic(
                            lines,
                            merge_cfg
                        )

            for idx, (ep, path) in enumerate(episodes.items(), 1):
                lines = get_lines_callback(ep)
                if not lines:
                    if progress_callback:
                        progress_callback(
                            idx,
                            total_episodes,
                            f"{translate_source('Пропуск серии')} {ep}..."
                        )
                    continue

                if progress_callback:
                    progress_callback(
                        idx - 1,
                        total_episodes,
                        f"{translate_source('Экспорт серии')} {ep}..."
                    )

                if do_html:
                    filename = f"{project_name} - Ep{ep}.html"
                    filepath = os.path.join(folder, filename)
                    html = self.generate_html(
                        ep,
                        self.process_merge_logic(lines, merge_cfg),
                        cfg,
                        cfg.get('highlight_ids_export'),
                        layout_type=cfg.get('layout_type', 'Таблица'),
                        is_editable=cfg.get('allow_edit', True)
                    )
                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write(html)
                    exported_count += 1

                if do_docx and DOCX_AVAILABLE:
                    filename = f"{project_name} - Ep{ep}.docx"
                    filepath = os.path.join(folder, filename)
                    document = self.create_docx_document(
                        {
                            ep: self.process_merge_logic(lines, merge_cfg)
                        },
                        cfg
                    )
                    document.save(filepath)
                    exported_count += 1

                if do_pdf:
                    filename = f"{project_name} - Ep{ep}.pdf"
                    filepath = os.path.join(folder, filename)
                    success, _ = self.export_to_pdf(
                        ep,
                        lines,
                        cfg,
                        filepath,
                        merge_cfg=merge_cfg
                    )
                    if success:
                        exported_count += 1

            if do_xls and EXCEL_AVAILABLE and all_episodes_data:
                filename = f"{project_name} - {translate_source('Все серии')}.xlsx"
                filepath = os.path.join(folder, filename)
                first_ep = next(iter(all_episodes_data))
                first_lines = all_episodes_data[first_ep]
                success, _ = self.export_to_excel(
                    first_ep, first_lines, cfg, filepath,
                    all_episodes_data, merge_cfg
                )
                if success:
                    exported_count += 1

            # Advance progress to completion
            if progress_callback:
                progress_callback(total_episodes, total_episodes, translate_source("Готово!"))

            # Open the folder
            if exported_count > 0 and cfg.get('open_auto', True):
                self._open_path(folder)

            return True, f"{translate_source('Экспортировано файлов:')} {exported_count}"

        except Exception as e:
            logger.error(f"Batch export error: {e}")
            return False, f"{translate_source('Ошибка пакетного экспорта:')} {e}"
