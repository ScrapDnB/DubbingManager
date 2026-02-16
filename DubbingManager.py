import sys
import json
import re
import os
import math
from datetime import datetime

# Импорты PySide6
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QHBoxLayout, QPushButton, QFileDialog, QTableWidget,
    QTableWidgetItem, QColorDialog, QComboBox, QLabel,
    QHeaderView, QInputDialog, QFrame, QSpinBox, QLineEdit,
    QDialog, QListWidget, QCheckBox, QGroupBox, QFormLayout,
    QMessageBox, QSlider, QAbstractItemView, QStackedWidget,
    QDoubleSpinBox, QRadioButton, QGridLayout, QScrollArea,
    QGraphicsView, QGraphicsScene, QGraphicsTextItem,
    QSplitter, QSizePolicy, QToolBar, QKeySequenceEdit, QDialogButtonBox
)
from PySide6.QtGui import (
    QColor, QFont, QPainter, QAction, QKeySequence, QPen, QBrush
)
from PySide6.QtCore import (
    Qt, QUrl, QTimer, QThread, Signal, QRectF, QEvent, Slot, QObject
)

# Для редактирования HTML
from PySide6.QtWebChannel import QWebChannel
from PySide6.QtCore import QObject, Slot

# Импорты для Мультимедиа
from PySide6.QtMultimedia import QMediaPlayer, QAudioOutput
from PySide6.QtMultimediaWidgets import QVideoWidget

# Импорт WebEngine
try:
    from PySide6.QtWebEngineWidgets import QWebEngineView
    WEB_ENGINE_AVAILABLE = True
except ImportError:
    from PySide6.QtWidgets import QTextBrowser as QWebEngineView
    WEB_ENGINE_AVAILABLE = False

# Попытка импорта openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Попытка импорта python-osc
try:
    from pythonosc.dispatcher import Dispatcher
    from pythonosc.osc_server import BlockingOSCUDPServer
    from pythonosc.udp_client import SimpleUDPClient
    OSC_AVAILABLE = True
except ImportError:
    OSC_AVAILABLE = False

# --- КОНСТАНТЫ ---
MY_PALETTE = [
    "#D9775F", "#E46C0A", "#9B5333", "#C0504D", "#C4BD97",
    "#D4A017", "#938953", "#8A7F80", "#76923C", "#4F6228",
    "#31859B", "#669999", "#4F81BD", "#5B9BD5", "#2C4D75",
    "#708090", "#B65C72", "#8064A2", "#5F497A", "#7B3F61"
]

# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---

def customize_table(table):
    """Настройка нативного вида таблиц"""
    table.setShowGrid(False)
    table.setAlternatingRowColors(True)
    table.setSelectionBehavior(QAbstractItemView.SelectRows)
    table.setSelectionMode(QAbstractItemView.ExtendedSelection)
    table.setFrameShape(QFrame.NoFrame)
    table.verticalHeader().setVisible(False)
    table.verticalHeader().setDefaultSectionSize(32)
    table.horizontalHeader().setHighlightSections(False)
    table.setStyleSheet("QTableWidget::item { padding-left: 10px; }")

def wrap_widget(widget):
    """Обертка для центрирования кнопок в таблице"""
    container = QWidget()
    layout = QHBoxLayout(container)
    layout.addWidget(widget)
    layout.setContentsMargins(4, 2, 4, 2)
    layout.setAlignment(Qt.AlignCenter)
    container.setLayout(layout)
    return container

def ass_time_to_seconds(time_str):
    try:
        h, m, s = time_str.split(':')
        return int(h) * 3600 + int(m) * 60 + float(s)
    except:
        return 0.0

def format_seconds_to_tc(seconds, round_flag=False):
    s = int(round(seconds)) if round_flag else int(seconds)
    hours = s // 3600
    minutes = (s % 3600) // 60
    secs = s % 60
    return f"{hours}:{minutes:02d}:{secs:02d}"

def hex_to_rgba_string(hex_code, alpha):
    """Преобразует HEX цвет в строку rgba(r, g, b, a)"""
    color = QColor(hex_code)
    if not color.isValid():
        return f"rgba(255, 255, 255, {alpha})"
    return f"rgba({color.red()}, {color.green()}, {color.blue()}, {alpha})"
    
class ReaperExportDialog(QDialog):
    def __init__(self, video_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки проекта Reaper")
        self.resize(350, 150)
        
        layout = QVBoxLayout(self)
        
        # Группировка опций
        layout.addWidget(QLabel("Выберите компоненты для экспорта:"))
        
        self.chk_video = QCheckBox("Добавить дорожку с видео")
        self.chk_regions = QCheckBox("Создать регионы (реплики с текстом)")
        
        # Настройка состояния чекбокса видео
        has_video = video_path and os.path.exists(video_path)
        if has_video:
            self.chk_video.setChecked(True)
            self.chk_video.setText(f"Добавить видео ({os.path.basename(video_path)})")
        else:
            self.chk_video.setChecked(False)
            self.chk_video.setEnabled(False)
            self.chk_video.setText("Видео не найдено (опция недоступна)")
            
        self.chk_regions.setChecked(True) # Регионы включены по умолчанию
        
        layout.addWidget(self.chk_video)
        layout.addWidget(self.chk_regions)
        
        # Кнопки ОК / Отмена
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def get_options(self):
        return self.chk_video.isChecked(), self.chk_regions.isChecked()

# --- WORKER ДЛЯ OSC (REAPER SYNC) ---
class OscWorker(QThread):
    time_changed = Signal(float)
    navigation_requested = Signal(str) # Передает 'next' или 'prev'
    
    def __init__(self, port=8000):
        super().__init__()
        self.port = port
        self.running = True
        self.server = None

    def run(self):
        if not OSC_AVAILABLE:
            return
        
        dispatcher = Dispatcher()
        
        # 1. Принимаем время (стандартные адреса Reaper)
        dispatcher.map("/time/seconds", self.handle_time)
        dispatcher.map("/time", self.handle_time)
        
        # 2. Принимаем навигацию через адрес имени трека (самый надежный способ обхода фильтров)
        # Мы слушаем изменение имени первого трека
        dispatcher.map("/track/1/name", self.handle_nav_via_name)
        
        # 3. Резервные прямые адреса
        dispatcher.map("/prompter/next", lambda addr, *args: self.navigation_requested.emit("next"))
        dispatcher.map("/prompter/prev", lambda addr, *args: self.navigation_requested.emit("prev"))
        
        # Логгер для отладки в консоли
        dispatcher.set_default_handler(self.debug_handler)

        try:
            self.server = BlockingOSCUDPServer(("127.0.0.1", self.port), dispatcher)
            self.server.timeout = 0.1
            while self.running:
                self.server.handle_request()
            self.server.server_close()
        except Exception as e:
            print(f"Ошибка OSC сервера: {e}")

    def debug_handler(self, address, *args):
        # Скрываем сообщения индикаторов громкости, чтобы не забивать консоль
        if "/vu" in address:
            return
        print(f"OSC Сообщение: {address} {args}")

    def handle_nav_via_name(self, address, *args):
        # Если в качестве имени трека пришло 'next' или 'prev'
        if args and isinstance(args[0], str):
            cmd = args[0].lower()
            if cmd == "next":
                print("ПОЛУЧЕНА КОМАНДА: СЛЕДУЮЩАЯ РЕПЛИКА")
                self.navigation_requested.emit("next")
            elif cmd == "prev":
                print("ПОЛУЧЕНА КОМАНДА: ПРЕДЫДУЩАЯ РЕПЛИКА")
                self.navigation_requested.emit("prev")

    def handle_time(self, address, *args):
        if args:
            try:
                self.time_changed.emit(float(args[0]))
            except:
                pass

    def stop(self):
        self.running = False

# --- МОСТ МЕЖДУ JS И PYTHON ---
class WebBridge(QObject):
    def __init__(self, main_app, parent=None):
        super().__init__(parent)
        self.main_app = main_app

    @Slot(int, int)
    def sync_scroll_index(self, index, total):
        """Обновляет счетчик в окне предпросмотра при прокрутке страницы"""
        if self.main_app.preview_window:
            self.main_app.preview_window.update_counter_label(index, total)

    @Slot(str, str)
    def update_text(self, line_id, new_text):
        """Принимает ID строки и новый текст из HTML"""
        try:
            lid = int(line_id)
            ep = self.main_app.ep_combo.currentData()
            
            # Ищем строку в памяти
            if ep in self.main_app.data.get("loaded_episodes", {}):
                lines = self.main_app.data["loaded_episodes"][ep]
                target = next((l for l in lines if l['id'] == lid), None)
                if target:
                    if target['text'] != new_text:
                        target['text'] = new_text
                        self.main_app.set_dirty(True)
                        # Отмечаем, что в окне предпросмотра есть реальные изменения текста
                        if self.main_app.preview_window:
                            self.main_app.preview_window._has_text_changes = True
                        print(f"Updated line {lid}: {new_text}")
        except Exception as e:
            print(f"Error updating text: {e}")

# --- ДИАЛОГ ФИЛЬТРА АКТЕРОВ ---
class ActorFilterDialog(QDialog):
    def __init__(self, actors_data, selected_ids, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выбор подсветки")
        self.resize(300, 400)
        self.actors_data = actors_data
        self.selected_ids = set(selected_ids) if selected_ids else set()
        
        layout = QVBoxLayout(self)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        self.checks_layout = QVBoxLayout(content)
        
        btn_box = QHBoxLayout()
        btn_all = QPushButton("Все")
        btn_none = QPushButton("Сбросить")
        btn_all.clicked.connect(self.select_all)
        btn_none.clicked.connect(self.select_none)
        btn_box.addWidget(btn_all)
        btn_box.addWidget(btn_none)
        layout.addLayout(btn_box)

        self.checkboxes = {}
        for aid, info in self.actors_data.items():
            chk = QCheckBox(info["name"])
            if not self.selected_ids or aid in self.selected_ids:
                chk.setChecked(True)
            self.checkboxes[aid] = chk
            self.checks_layout.addWidget(chk)
            
        self.checks_layout.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        btn_ok = QPushButton("Применить")
        btn_ok.clicked.connect(self.accept)
        layout.addWidget(btn_ok)

    def select_all(self):
        for chk in self.checkboxes.values():
            chk.setChecked(True)

    def select_none(self):
        for chk in self.checkboxes.values():
            chk.setChecked(False)

    def get_selected(self):
        return [aid for aid, chk in self.checkboxes.items() if chk.isChecked()]
        
# --- ДИАЛОГ НАСТРОЙКИ ЦВЕТОВ ---
class PrompterColorDialog(QDialog):
    def __init__(self, current_colors, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройка цветовой схемы телесуфлёра")
        self.resize(450, 400)
        self.colors = current_colors.copy()
        
        main_layout = QVBoxLayout(self)
        form_layout = QFormLayout()

        self.btns = {}
        # Полные названия всех элементов для настройки
        self.color_names = {
            "bg": "Фоновый цвет сцены суфлёра",
            "active_text": "Цвет текста активной реплики",
            "inactive_text": "Цвет текста неактивной реплики",
            "tc": "Цвет таймкода внутри реплики",
            "actor": "Цвет имени актёра в реплике",
            "header_bg": "Фон верхней панели таймкода",
            "header_text": "Цвет цифр таймкода Reaper"
        }

        for key, display_name in self.color_names.items():
            btn = QPushButton()
            btn.setFixedSize(80, 25)
            # Отображаем текущий цвет на самой кнопке
            btn.setStyleSheet(f"background-color: {self.colors[key]}; border: 1px solid #555; border-radius: 4px;")
            btn.clicked.connect(lambda checked=False, k=key: self.pick_color_for(k))
            self.btns[key] = btn
            form_layout.addRow(display_name, btn)

        main_layout.addLayout(form_layout)

        # Нижние кнопки диалога
        dialog_buttons = QHBoxLayout()
        btn_save = QPushButton("Сохранить цветовую схему")
        btn_save.clicked.connect(self.accept)
        btn_cancel = QPushButton("Отмена")
        btn_cancel.clicked.connect(self.reject)
        
        dialog_buttons.addWidget(btn_save)
        dialog_buttons.addWidget(btn_cancel)
        main_layout.addLayout(dialog_buttons)

    def pick_color_for(self, key):
        initial = QColor(self.colors[key])
        new_color = QColorDialog.getColor(initial, self, "Выберите цвет")
        if new_color.isValid():
            hex_val = new_color.name()
            self.colors[key] = hex_val
            self.btns[key].setStyleSheet(f"background-color: {hex_val}; border: 1px solid #555; border-radius: 4px;")

    def get_final_colors(self):
        return self.colors
        
# --- ОКНО ТЕЛЕСУФЛЁРА ---
class TeleprompterWindow(QDialog):
    def __init__(self, main_app, ep_num):
        super().__init__(main_app)
        self.main_app = main_app
        self.ep_num = ep_num
        self.setWindowTitle(f"Телесуфлёр - Серия {ep_num}")
        self.resize(1200, 900)
        
        # --- ГЛУБОКАЯ ИНИЦИАЛИЗАЦИЯ НАСТРОЕК (ЗАЩИТА ОТ ПАДЕНИЙ ПРИ ОТКРЫТИИ) ---
        default_config = {
            "f_tc": 20, "f_char": 24, "f_actor": 18, "f_text": 36,
            "focus_ratio": 0.5, "is_mirrored": False, "show_header": False,
            "port_in": 8000, "port_out": 9000,
            "sync_in": True, "sync_out": False,
            "key_prev": "Left", "key_next": "Right",
            # slider value 0..100; 0 = instant, higher = longer smoothing up to ~2s
            "scroll_smoothness_slider": 18,
            "colors": {
                "bg": "#000000",
                "active_text": "#FFFFFF",
                "inactive_text": "#444444",
                "tc": "#888888",
                "actor": "#AAAAAA",
                "header_bg": "#111111",
                "header_text": "#00FF00"
            }
        }

        # Проверяем, есть ли секция настроек в проекте
        if "prompter_config" not in self.main_app.data or self.main_app.data["prompter_config"] is None:
            self.main_app.data["prompter_config"] = default_config
        
        self.cfg = self.main_app.data["prompter_config"]

        # Проверяем наличие словаря цветов (для совместимости со старыми сохранениями)
        if "colors" not in self.cfg or not isinstance(self.cfg["colors"], dict):
            self.cfg["colors"] = default_config["colors"]
        else:
            # Если каких-то отдельных ключей цвета не хватает
            for color_key, color_value in default_config["colors"].items():
                if color_key not in self.cfg["colors"]:
                    self.cfg["colors"][color_key] = color_value
        
        # Переменные состояния
        self.time_map = []
        self.osc_thread = None
        self.osc_client = None
        self.last_known_time = 0.0
        self.highlight_ids = None

        # Инициализация интерфейса
        self.init_ui()
        # Построение графических объектов
        self.build_prompter_content()

    def init_ui(self):
        # Основной вертикальный слой
        self.root_layout = QVBoxLayout(self)
        self.root_layout.setContentsMargins(0, 0, 0, 0)
        self.root_layout.setSpacing(0)

        # --- ВЕРХНИЙ ТУЛБАР (Навигация) ---
        self.toolbar = QToolBar("Управление")
        self.toolbar.setMovable(False)
        self.toolbar.setStyleSheet("QToolBar { padding: 5px; background: #333; border-bottom: 1px solid #111; }")

        self.btn_toggle_settings = QPushButton("⚙ Панель настроек")
        self.btn_toggle_settings.setCheckable(True)
        self.btn_toggle_settings.clicked.connect(self.toggle_settings_panel_visibility)
        self.toolbar.addWidget(self.btn_toggle_settings)

        self.toolbar.addSeparator()

        self.btn_go_prev = QPushButton("⏮ Предыдущая реплика")
        self.btn_go_prev.setMinimumWidth(160)
        self.btn_go_prev.clicked.connect(lambda: self.navigate_to_replica_in_direction(-1))
        self.toolbar.addWidget(self.btn_go_prev)

        self.btn_go_next = QPushButton("Следующая реплика ⏭")
        self.btn_go_next.setMinimumWidth(160)
        self.btn_go_next.clicked.connect(lambda: self.navigate_to_replica_in_direction(1))
        self.toolbar.addWidget(self.btn_go_next)

        # Распорка для выравнивания
        toolbar_spacer = QWidget()
        toolbar_spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.toolbar.addWidget(toolbar_spacer)
        
        btn_close_window = QPushButton("Закрыть окно")
        btn_close_window.clicked.connect(self.close)
        self.toolbar.addWidget(btn_close_window)

        self.root_layout.addWidget(self.toolbar)

        # --- ВЕРТИКАЛЬНЫЙ СПЛИТТЕР ---
        self.v_splitter = QSplitter(Qt.Vertical)
        self.v_splitter.setHandleWidth(8)
        self.v_splitter.setStyleSheet("QSplitter::handle { background: #444; }")
        self.v_splitter.splitterMoved.connect(lambda pos, idx: self.update_big_timecode_font_size())

        # Верхняя панель большого таймкода
        self.header_panel = QFrame()
        self.header_panel.setObjectName("HeaderPanel")
        self.header_panel.setMinimumHeight(0)
        header_layout = QVBoxLayout(self.header_panel)
        header_layout.setContentsMargins(0, 0, 0, 0)
        
        self.lbl_big_timecode = QLabel("0:00:00.000")
        self.lbl_big_timecode.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(self.lbl_big_timecode)
        self.v_splitter.addWidget(self.header_panel)

        # --- ГОРИЗОНТАЛЬНЫЙ СПЛИТТЕР ---
        self.h_splitter = QSplitter(Qt.Horizontal)
        self.h_splitter.setHandleWidth(8)

        # ЛЕВАЯ ПАНЕЛЬ (Настройки и список)
        self.side_panel_widget = QWidget()
        self.side_panel_widget.setMinimumWidth(320)
        self.side_layout = QVBoxLayout(self.side_panel_widget)

        # Зона настроек с прокруткой
        self.settings_scroll_area = QScrollArea()
        self.settings_scroll_area.setWidgetResizable(True)
        self.settings_scroll_area.setFrameShape(QFrame.NoFrame)
        
        settings_container = QWidget()
        settings_v_layout = QVBoxLayout(settings_container)

        # 1. Настройка шрифтов
        fonts_group_box = QGroupBox("Размеры шрифтов элементов")
        fonts_form = QFormLayout(fonts_group_box)
        self.spin_font_tc = QSpinBox(); self.spin_font_tc.setRange(10, 150); self.spin_font_tc.setValue(self.cfg["f_tc"])
        self.spin_font_char = QSpinBox(); self.spin_font_char.setRange(10, 150); self.spin_font_char.setValue(self.cfg["f_char"])
        self.spin_font_actor = QSpinBox(); self.spin_font_actor.setRange(10, 150); self.spin_font_actor.setValue(self.cfg["f_actor"])
        self.spin_font_text = QSpinBox(); self.spin_font_text.setRange(10, 300); self.spin_font_text.setValue(self.cfg["f_text"])
        
        for s in [self.spin_font_tc, self.spin_font_char, self.spin_font_actor, self.spin_font_text]:
            s.valueChanged.connect(self.handle_font_config_change)
        
        fonts_form.addRow("Таймкод:", self.spin_font_tc)
        fonts_form.addRow("Имя персонажа:", self.spin_font_char)
        fonts_form.addRow("Имя актёра:", self.spin_font_actor)
        fonts_form.addRow("Текст реплики:", self.spin_font_text)
        settings_v_layout.addWidget(fonts_group_box)

        # 2. Настройка Фокуса
        focus_group_box = QGroupBox("Позиция линии чтения")
        focus_layout = QVBoxLayout(focus_group_box)
        self.slider_focus_pos = QSlider(Qt.Horizontal); self.slider_focus_pos.setRange(10, 90)
        self.slider_focus_pos.setValue(int(self.cfg["focus_ratio"] * 100))
        self.slider_focus_pos.valueChanged.connect(self.handle_focus_ratio_change)
        
        self.lbl_focus_percent = QLabel(f"Высота линии: {self.slider_focus_pos.value()}%")
        self.lbl_focus_percent.setAlignment(Qt.AlignCenter)
        focus_layout.addWidget(self.lbl_focus_percent)
        focus_layout.addWidget(self.slider_focus_pos)
        settings_v_layout.addWidget(focus_group_box)

        # 4a. Плавность прокрутки (ползунок)
        scroll_group = QGroupBox("Прокрутка")
        # Используем вертикальный лэйаут, как в блоке "Позиция линии чтения"
        sg_layout = QVBoxLayout(scroll_group)
        # Используем слайдер 0..90, отображаемое значение 0.00..0.90
        self.slider_scroll_smoothness = QSlider(Qt.Horizontal)
        self.slider_scroll_smoothness.setRange(0, 100)
        # Initialize slider from new or legacy config
        if "scroll_smoothness_slider" in self.cfg:
            init_val = int(self.cfg.get("scroll_smoothness_slider", 18))
        else:
            # legacy float value (0.0..0.9) -> convert
            init_val = int(round(self.cfg.get("scroll_smoothness", 0.18) * 100))
        self.slider_scroll_smoothness.setValue(init_val)
        self.slider_scroll_smoothness.valueChanged.connect(self.handle_scroll_smoothness_change)
        # Метка для отображения текущего значения как дроби
        self.lbl_scroll_value = QLabel(f"{self.cfg.get('scroll_smoothness', 0.18):.2f}")
        # Метка описания над слайдером
        self.lbl_scroll_descr = QLabel("Плавность прокрутки (слева — быстрее, справа — дольше задержка):")
        self.lbl_scroll_descr.setAlignment(Qt.AlignCenter)
        sg_layout.addWidget(self.lbl_scroll_descr)
        # Горизонтальная строка для слайдера + метки текущего значения
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.addWidget(self.slider_scroll_smoothness)
        row_layout.addWidget(self.lbl_scroll_value)
        sg_layout.addWidget(row_widget)
        settings_v_layout.addWidget(scroll_group)
        # Обновим начальное отображение метки исходя из текущего положения слайдера
        self.handle_scroll_smoothness_change()

        # 3. Вид и Цвета
        view_group = QGroupBox("Отображение")
        view_lay = QVBoxLayout(view_group)
        btn_open_colors = QPushButton("🎨 Настроить цвета телесуфлёра...")
        btn_open_colors.clicked.connect(self.open_color_settings_dialog)
        self.chk_show_header_panel = QCheckBox("Показать таймкод Reaper сверху", checked=self.cfg["show_header"])
        self.chk_show_header_panel.toggled.connect(self.toggle_header_visibility)
        self.chk_mirror_mode_active = QCheckBox("Зеркальный режим (отражение)", checked=self.cfg.get("is_mirrored", False))
        self.chk_mirror_mode_active.toggled.connect(self.toggle_mirror_mode)
        view_lay.addWidget(btn_open_colors); view_lay.addWidget(self.chk_show_header_panel); view_lay.addWidget(self.chk_mirror_mode_active)
        settings_v_layout.addWidget(view_group)

        # 4. Синхронизация Reaper
        osc_group_box = QGroupBox("Синхронизация Reaper (OSC)")
        osc_layout = QVBoxLayout(osc_group_box)
        self.chk_follow_reaper_in = QCheckBox("Суфлёр следует за Reaper", checked=self.cfg["sync_in"])
        self.chk_reaper_follow_out = QCheckBox("Reaper следует за навигацией", checked=self.cfg["sync_out"])
        self.chk_follow_reaper_in.toggled.connect(self.save_current_config_to_project)
        self.chk_reaper_follow_out.toggled.connect(self.save_current_config_to_project)
        self.btn_activate_osc_link = QPushButton("Включить OSC связь")
        self.btn_activate_osc_link.setCheckable(True)
        self.btn_activate_osc_link.clicked.connect(self.toggle_osc_connection_status)
        osc_layout.addWidget(self.chk_follow_reaper_in); osc_layout.addWidget(self.chk_reaper_follow_out); osc_layout.addWidget(self.btn_activate_osc_link)
        settings_v_layout.addWidget(osc_group_box)

        self.settings_scroll_area.setWidget(settings_container)
        self.side_layout.addWidget(self.settings_scroll_area)

        # --- СПИСОК РЕПЛИК (АДАПТИВНЫЙ) ---
        self.side_layout.addWidget(QLabel("<b>Список реплик актёра:</b>"))
        self.list_of_replicas = QListWidget()
        self.list_of_replicas.itemClicked.connect(self.on_replica_list_item_clicked)
        # Заставляем список занимать всё доступное место
        self.list_of_replicas.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        self.side_layout.addWidget(self.list_of_replicas)

        self.btn_actor_filter = QPushButton("Выбор актёров для суфлёра..."); self.btn_actor_filter.clicked.connect(self.open_actor_filter_dialog)
        self.side_layout.addWidget(self.btn_actor_filter)

        self.h_splitter.addWidget(self.side_panel_widget)

        # --- ГРАФИЧЕСКАЯ СЦЕНА ---
        self.prompter_scene = QGraphicsScene()
        self.prompter_view = QGraphicsView(self.prompter_scene)
        self.prompter_view.setRenderHints(QPainter.Antialiasing | QPainter.TextAntialiasing)
        self.prompter_view.setDragMode(QGraphicsView.ScrollHandDrag)
        self.prompter_view.setFrameShape(QFrame.NoFrame)
        self.prompter_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.prompter_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.h_splitter.addWidget(self.prompter_view)
        
        self.v_splitter.addWidget(self.h_splitter)
        self.root_layout.addWidget(self.v_splitter)

        # Инициализация видимости хедера
        self.header_panel.setVisible(self.cfg["show_header"])
        self.v_splitter.setSizes([100, 800])
        self.h_splitter.setSizes([320, 900])

        # Таймер плавной прокрутки и целевая позиция
        self.smooth_scroll_timer = QTimer()
        self.smooth_scroll_timer.setInterval(16)  # ~60 FPS
        self.smooth_scroll_timer.timeout.connect(self.smooth_scroll_step)
        self._scroll_target_y = None

    def compute_scroll_tau(self):
        """Возвращает временную константу (tau) в секундах, соответствующую положению слайдера.
        Slider: 0..100 -> tau: 0 (instant) .. max_tau (slower, holds up to ~2s).
        Нелинейное отображение даёт удобную чувствительность слева и большие времена справа.
        """
        s = None
        if hasattr(self, 'slider_scroll_smoothness'):
            s = int(self.slider_scroll_smoothness.value())
        else:
            s = int(self.cfg.get('scroll_smoothness_slider', 18))

        if s <= 0:
            return 0.0

        # mapping parameters
        min_tau = 0.01
        max_tau = 2.0
        p = 1.15
        n = float(s) / 100.0
        tau = min_tau + (n ** p) * (max_tau - min_tau)
        return tau

    # --- ЛОГИКА ---

    def update_big_timecode_font_size(self):
        """Динамический пересчет шрифта при изменении размеров сплиттера"""
        current_h = self.header_panel.height()
        if current_h > 10:
            font_size = int(current_h * 0.7)
            text_color = self.cfg['colors']['header_text']
            self.lbl_big_timecode.setStyleSheet(
                f"color: {text_color}; font-family: 'Courier New'; font-weight: bold; font-size: {font_size}px;"
            )

    def toggle_settings_panel_visibility(self, is_hidden):
        """Скрытие левой панели настроек"""
        self.side_panel_widget.setVisible(not is_hidden)
        if is_hidden:
            self.btn_toggle_settings.setText("⚙ Показать настройки")
        else:
            self.btn_toggle_settings.setText("⚙ Скрыть настройки")

    def handle_font_config_change(self):
        """Сохранение размеров шрифтов и перерисовка сцены"""
        self.cfg["f_tc"] = self.spin_font_tc.value()
        self.cfg["f_char"] = self.spin_font_char.value()
        self.cfg["f_actor"] = self.spin_font_actor.value()
        self.cfg["f_text"] = self.spin_font_text.value()
        self.main_app.set_dirty(True)
        self.build_prompter_content()

    def handle_focus_ratio_change(self):
        """Изменение точки фокуса чтения"""
        val = self.slider_focus_pos.value()
        self.cfg["focus_ratio"] = val / 100.0
        self.lbl_focus_percent.setText(f"Высота линии: {val}%")
        self.main_app.set_dirty(True)
        self.update_view_position_by_time(self.last_known_time)

    def handle_scroll_smoothness_change(self):
        """Сохранение параметра плавности прокрутки"""
        # slider value is 0..100; store slider integer and show mapped tau
        sval = int(self.slider_scroll_smoothness.value())
        self.cfg["scroll_smoothness_slider"] = sval
        tau = self.compute_scroll_tau()
        if tau <= 0:
            self.lbl_scroll_value.setText("instant")
            self.lbl_scroll_descr.setText("Плавность прокрутки: мгновенно (без сглаживания)")
        else:
            # show tau in seconds with 2 decimals
            self.lbl_scroll_value.setText(f"{tau:.2f}s")
            self.lbl_scroll_descr.setText(f"Плавность прокрутки: задержка ≈ {tau:.2f}s")
        self.main_app.set_dirty(True)

    def smooth_scroll_step(self):
        """Выполняет шаг интерполяции центра вида к целевой Y-позиции"""
        if self._scroll_target_y is None:
            self.smooth_scroll_timer.stop()
            return

        view = self.prompter_view
        vp_center = view.viewport().rect().center()
        scene_center = view.mapToScene(vp_center)
        current_y = scene_center.y()
        target_y = self._scroll_target_y

        # Compute time-constant tau from slider and convert to per-tick alpha
        tau = self.compute_scroll_tau()
        if tau <= 0:
            view.centerOn(425, target_y)
            self._scroll_target_y = None
            self.smooth_scroll_timer.stop()
            return

        dt = max(0.001, float(self.smooth_scroll_timer.interval()) / 1000.0)
        alpha = 1.0 - math.exp(-dt / tau)
        new_y = current_y * (1.0 - alpha) + target_y * alpha

        if abs(new_y - target_y) < 0.5:
            view.centerOn(425, target_y)
            self._scroll_target_y = None
            self.smooth_scroll_timer.stop()
        else:
            view.centerOn(425, new_y)

    def open_color_settings_dialog(self):
        """Открытие диалога настройки цветов"""
        dialog = PrompterColorDialog(self.cfg["colors"], self)
        if dialog.exec():
            self.cfg["colors"] = dialog.get_final_colors()
            self.main_app.set_dirty(True)
            self.build_prompter_content()

    def build_prompter_content(self):
        """Построение всех текстовых блоков на графической сцене"""
        clrs = self.cfg["colors"]
        self.prompter_scene.clear()
        self.time_map = []
        self.list_of_replicas.clear()
        
        # Применяем цвета фона
        self.prompter_scene.setBackgroundBrush(QColor(clrs["bg"]))
        self.header_panel.setStyleSheet(f"background-color: {clrs['header_bg']};")
        self.update_big_timecode_font_size()

        lines = self.main_app.get_episode_lines(self.ep_num)
        if not lines:
            return
        
        lines.sort(key=lambda x: x['s'])
        processed_lines = self.main_app.process_merge_logic(lines, self.main_app.data["export_config"])
        
        y_cursor = 1000.0 # Начальный отступ
        width = 850
        
        # Шрифты для отрисовки элементов
        f_tc = QFont("Courier New", self.cfg["f_tc"])
        f_char = QFont("Arial", self.cfg["f_char"], QFont.Bold)
        f_actor = QFont("Arial", self.cfg["f_actor"]); f_actor.setItalic(True)
        f_text = QFont("Arial", self.cfg["f_text"])
        
        # Фильтр актёров
        all_actor_ids = set(self.main_app.data["actors"].keys())
        active_actor_ids = set(self.highlight_ids) if self.highlight_ids is not None else all_actor_ids

        for i, replica in enumerate(processed_lines):
            actor_id = self.main_app.data["global_map"].get(replica['char'])
            actor_info = self.main_app.data["actors"].get(actor_id, {"name": "-", "color": "#FFFFFF"})
            is_replica_active = actor_id in active_actor_ids
            
            # Применение цветов из конфига
            if is_replica_active:
                char_col = QColor(actor_info['color'])
                if char_col.value() < 100: char_col = QColor("white")
                text_col = QColor(clrs["active_text"])
                tc_col = QColor(clrs["tc"])
                # Добавление в список навигации
                self.list_of_replicas.addItem(f"{format_seconds_to_tc(replica['s'])} - {replica['char']}")
                self.list_of_replicas.item(self.list_of_replicas.count()-1).setData(Qt.UserRole, replica['s'])
            else:
                inactive_col = QColor(clrs["inactive_text"])
                char_col = text_col = tc_col = inactive_col

            row_y = y_cursor
            # Отрисовка Имени персонажа
            item_char = QGraphicsTextItem(replica['char']); item_char.setFont(f_char); item_char.setDefaultTextColor(char_col)
            item_char.setPos(0, row_y); self.prompter_scene.addItem(item_char)
            
            # Отрисовка Таймкода
            item_tc = QGraphicsTextItem(f"[{format_seconds_to_tc(replica['s'])}]"); item_tc.setFont(f_tc); item_tc.setDefaultTextColor(tc_col)
            item_tc.setPos(item_char.boundingRect().width() + 20, row_y + (self.cfg["f_char"] - self.cfg["f_tc"]) / 2); self.prompter_scene.addItem(item_tc)
            
            # Отрисовка Имени актёра
            item_actor = QGraphicsTextItem(f"({actor_info['name']})"); item_actor.setFont(f_actor); item_actor.setDefaultTextColor(char_col)
            item_actor.setPos(item_char.boundingRect().width() + item_tc.boundingRect().width() + 40, row_y + (self.cfg["f_char"] - self.cfg["f_actor"]) / 2); self.prompter_scene.addItem(item_actor)
            
            y_cursor += item_char.boundingRect().height()
            
            # Отрисовка Текста реплики
            item_main_text = QGraphicsTextItem(replica['text']); item_main_text.setFont(f_text); item_main_text.setDefaultTextColor(text_col)
            item_main_text.setTextWidth(width); item_main_text.setPos(0, y_cursor); self.prompter_scene.addItem(item_main_text)
            
            # Сохраняем маппинг центра блока для скролла
            self.time_map.append({
                'index': i, 's': replica['s'], 'e': replica['e'],
                'y_center': row_y + (item_char.boundingRect().height() + item_main_text.boundingRect().height()) / 2,
                'active': is_replica_active
            })
            y_cursor += item_main_text.boundingRect().height() + (self.cfg["f_text"] * 1.8)

        self.prompter_scene.setSceneRect(-50, 0, width + 100, y_cursor + 1000)
        self.update_view_position_by_time(self.last_known_time)

    def update_view_position_by_time(self, time_seconds):
        """Главный метод синхронизации всех элементов окна"""
        self.last_known_time = time_seconds
        
        # 1. Таймкод Reaper
        ms = int((time_seconds % 1) * 1000)
        self.lbl_big_timecode.setText(f"{format_seconds_to_tc(time_seconds)}.{ms:03d}")
        if self.header_panel.isVisible():
            self.update_big_timecode_font_size()

        if not self.time_map:
            return

        target_y_coordinate = 0
        target_list_item_index = -1
        
        # 2. Поиск позиции на сцене и в списке
        for i, segment in enumerate(self.time_map):
            if time_seconds >= segment['s'] and time_seconds <= segment['e']:
                target_y_coordinate = segment['y_center']
                if segment['active']:
                    for row in range(self.list_of_replicas.count()):
                        if abs(self.list_of_replicas.item(row).data(Qt.UserRole) - segment['s']) < 0.01:
                            target_list_item_index = row; break
                break
            elif time_seconds < segment['s']:
                prev = self.time_map[i-1] if i > 0 else None
                if prev:
                    ratio = (time_seconds - prev['e']) / (segment['s'] - prev['e']) if (segment['s'] - prev['e']) > 0 else 0
                    target_y_coordinate = prev['y_center'] + (segment['y_center'] - prev['y_center']) * ratio
                else:
                    target_y_coordinate = segment['y_center']
                break
            if i == len(self.time_map) - 1:
                target_y_coordinate = segment['y_center'] + (time_seconds - segment['e']) * 100

        # 3. Синхронизация бокового списка (слежение за текущей репликой)
        if target_list_item_index != -1:
            self.list_of_replicas.blockSignals(True)
            self.list_of_replicas.setCurrentRow(target_list_item_index)
            # Плавная прокрутка списка так, чтобы текущий элемент был в центре
            self.list_of_replicas.scrollToItem(self.list_of_replicas.currentItem(), QAbstractItemView.PositionAtCenter)
            self.list_of_replicas.blockSignals(False)

        # 4. Прокрутка сцены суфлёра
        view_h = self.prompter_view.height()
        offset = (0.5 - self.cfg["focus_ratio"]) * view_h
        # Плавная прокрутка: устанавливаем целевую позицию и запускаем таймер
        target_full_y = target_y_coordinate + offset
        # Use slider-derived tau for deterministic smoothing behavior
        tau = self.compute_scroll_tau()
        if tau <= 0.0:
            # Мгновенное позиционирование
            self._scroll_target_y = None
            self.prompter_view.centerOn(425, target_full_y)
            if self.smooth_scroll_timer.isActive():
                self.smooth_scroll_timer.stop()
        else:
            self._scroll_target_y = target_full_y
            if not self.smooth_scroll_timer.isActive():
                self.smooth_scroll_timer.start()

    def jump_to_specific_time(self, t):
        """Метод для мгновенного прыжка к таймкоду"""
        self.last_known_time = t
        self.update_view_position_by_time(t)
        
        # Отправка OSC обратно в Reaper (прыжок курсора)
        if OSC_AVAILABLE and self.btn_activate_osc_link.isChecked() and self.cfg["sync_out"]:
            if self.osc_client:
                try:
                    self.osc_client.send_message("/time", float(t))
                    self.osc_client.send_message("/track/0/pos", float(t))
                except:
                    pass

    def navigate_to_replica_in_direction(self, direction):
        """Перемещение по активным репликам (назад/вперед)"""
        if not self.time_map: return
        curr_idx = 0; min_d = 999999
        for seg in self.time_map:
            d = abs(self.last_known_time - seg['s'])
            if d < min_d: min_d = d; curr_idx = seg['index']
        
        target_idx = curr_idx
        while True:
            target_idx += direction
            if 0 <= target_idx < len(self.time_map):
                if self.time_map[target_idx]['active']:
                    self.jump_to_specific_time(self.time_map[target_idx]['s'])
                    break
            else:
                break

    def on_replica_list_item_clicked(self, item):
        self.jump_to_specific_time(item.data(Qt.UserRole))

    def save_current_config_to_project(self):
        """Сохранение галочек синхронизации в данные проекта"""
        self.cfg["sync_in"] = self.chk_follow_reaper_in.isChecked()
        self.cfg["sync_out"] = self.chk_reaper_follow_out.isChecked()
        self.main_app.set_dirty(True)

    def toggle_osc_connection_status(self, checked):
        """Включение и выключение OSC"""
        if checked:
            self.osc_thread = OscWorker(self.cfg["port_in"])
            self.osc_thread.time_changed.connect(self.on_osc_time_packet_received)
            # Слушаем команды навигации из Lua скриптов
            self.osc_thread.navigation_requested.connect(self.navigate_to_replica_in_direction_from_osc)
            self.osc_thread.start()
            try:
                self.osc_client = SimpleUDPClient("127.0.0.1", self.cfg["port_out"])
                self.btn_activate_osc_link.setText("OSC Связь: Активна")
            except:
                self.btn_activate_osc_link.setText("Ошибка OSC")
        else:
            if self.osc_thread:
                self.osc_thread.stop()
                self.osc_thread = None
            self.osc_client = None
            self.btn_activate_osc_link.setText("Включить OSC связь")
            self.btn_activate_osc_link.setStyleSheet("")

    @Slot(float)
    def on_osc_time_packet_received(self, time_val):
        if self.cfg["sync_in"]:
            self.update_view_position_by_time(time_val)

    @Slot(str)
    def navigate_to_replica_in_direction_from_osc(self, direction):
        step = 1 if direction == "next" else -1
        self.navigate_to_replica_in_direction(step)

    def toggle_mirror_mode(self, checked):
        self.cfg["is_mirrored"] = checked
        self.prompter_view.resetTransform()
        if checked:
            self.prompter_view.scale(-1, 1)
        self.main_app.set_dirty(True)

    def toggle_header_visibility(self, checked):
        self.cfg["show_header"] = checked
        self.header_panel.setVisible(checked)
        QTimer.singleShot(50, self.update_big_timecode_font_size)
        self.main_app.set_dirty(True)

    def open_actor_filter_dialog(self):
        all_ids = list(self.main_app.data["actors"].keys())
        d = ActorFilterDialog(self.main_app.data["actors"], self.highlight_ids or all_ids, self)
        if d.exec():
            sel = d.get_selected()
            self.highlight_ids = None if len(sel) == len(all_ids) or len(sel) == 0 else sel
            self.build_prompter_content()

    def keyPressEvent(self, event):
        # Навигация клавишами, если окно активно
        if event.key() == Qt.Key_Left: self.navigate_to_replica_in_direction(-1)
        elif event.key() == Qt.Key_Right: self.navigate_to_replica_in_direction(1)
        else: super().keyPressEvent(event)

    def closeEvent(self, event):
        if self.osc_thread:
            self.osc_thread.stop()
        event.accept()

# --- ОКНО ЖИВОГО ПРЕДПРОСМОТРА HTML ---
class HtmlLivePreview(QDialog):
    def __init__(self, main_app, ep_num):
        super().__init__(main_app)
        self.main_app = main_app
        self.ep_num = ep_num
        self.setWindowTitle(f"Предпросмотр монтажного листа: Серия {ep_num}")
        self.resize(1200, 900)
        
        self.highlight_ids = None
        self.current_h_index = -1
        self._has_text_changes = False  # Флаг для отслеживания реальных изменений текста
        
        self.init_ui()
        
        self.browser.loadFinished.connect(self.on_page_loaded)
        
        # Убедимся что главное приложение знает о нас
        self.main_app.preview_window = self
        
        if WEB_ENGINE_AVAILABLE:
            self.channel = QWebChannel()
            self.bridge = WebBridge(self.main_app)
            self.channel.registerObject("backend", self.bridge)
            self.browser.page().setWebChannel(self.channel)

        self.update_preview()

    def init_ui(self):
        self.root_layout = QVBoxLayout(self)
        
        self.nav_panel = QHBoxLayout()
        self.btn_toggle_sidebar = QPushButton("⬅ Скрыть настройки")
        self.btn_toggle_sidebar.setCheckable(True)
        self.btn_toggle_sidebar.clicked.connect(self.toggle_sidebar)
        
        self.btn_prev_h = QPushButton("⏮ Пред. реплика (Alt+←)")
        self.btn_prev_h.setShortcut("Alt+Left")
        self.btn_prev_h.clicked.connect(lambda: self.scroll_to_highlight("prev"))
        
        self.lbl_h_count = QLabel("0 / 0")
        self.lbl_h_count.setStyleSheet("font-weight: bold; margin: 0 10px;")
        
        self.btn_next_h = QPushButton("След. реплика (Alt+→) ⏭")
        self.btn_next_h.setShortcut("Alt+Right")
        self.btn_next_h.clicked.connect(lambda: self.scroll_to_highlight("next"))
        
        self.nav_panel.addWidget(self.btn_toggle_sidebar)
        self.nav_panel.addSpacing(20)
        self.nav_panel.addWidget(self.btn_prev_h)
        self.nav_panel.addWidget(self.lbl_h_count)
        self.nav_panel.addWidget(self.btn_next_h)
        self.nav_panel.addStretch()
        
        self.root_layout.addLayout(self.nav_panel)

        self.content_layout = QHBoxLayout()
        self.root_layout.addLayout(self.content_layout)
        
        self.settings_panel = QFrame()
        self.settings_panel.setFixedWidth(280)
        self.settings_panel.setFrameShape(QFrame.StyledPanel)
        sp_layout = QVBoxLayout(self.settings_panel)
        
        sp_layout.addWidget(QLabel("<b>Настройки вида</b>"))
        
        self.combo_layout = QComboBox()
        self.combo_layout.addItems(["Таблица", "Сценарий"])
        current_type = self.main_app.data["export_config"].get("layout_type", "Таблица")
        self.combo_layout.setCurrentText(current_type)
        self.combo_layout.currentIndexChanged.connect(self.update_preview)
        sp_layout.addWidget(QLabel("Формат:"))
        sp_layout.addWidget(self.combo_layout)
        sp_layout.addSpacing(10)

        font_group = QGroupBox("Размеры шрифтов")
        fg_layout = QFormLayout(font_group)
        self.s_time = QSpinBox(); self.s_time.setRange(6, 48)
        self.s_time.setValue(self.main_app.data["export_config"].get("f_time", 12))
        self.s_time.valueChanged.connect(self.on_setting_change)
        
        self.s_char = QSpinBox(); self.s_char.setRange(6, 48)
        self.s_char.setValue(self.main_app.data["export_config"].get("f_char", 14))
        self.s_char.valueChanged.connect(self.on_setting_change)
        
        self.s_actor = QSpinBox(); self.s_actor.setRange(6, 48)
        self.s_actor.setValue(self.main_app.data["export_config"].get("f_actor", 14))
        self.s_actor.valueChanged.connect(self.on_setting_change)
        
        self.s_text = QSpinBox(); self.s_text.setRange(6, 48)
        self.s_text.setValue(self.main_app.data["export_config"].get("f_text", 16))
        self.s_text.valueChanged.connect(self.on_setting_change)
        
        fg_layout.addRow("Таймкод:", self.s_time)
        fg_layout.addRow("Персонаж:", self.s_char)
        fg_layout.addRow("Актер:", self.s_actor)
        fg_layout.addRow("Текст:", self.s_text)
        sp_layout.addWidget(font_group)
        
        filter_group = QGroupBox("Подсветка")
        f_lay = QVBoxLayout(filter_group)
        btn_filter = QPushButton("Выбрать актеров...")
        btn_filter.clicked.connect(self.open_actor_filter)
        f_lay.addWidget(btn_filter)
        sp_layout.addWidget(filter_group)
        sp_layout.addStretch()
        
        save_group = QGroupBox("Сохранение")
        sg_layout = QVBoxLayout(save_group)
        btn_save_ass = QPushButton("💾 Сохранить в .ASS")
        btn_save_ass.clicked.connect(self.save_to_original_ass)
        btn_save_copy = QPushButton("Сохранить копию...")
        btn_save_copy.clicked.connect(self.save_ass_copy)
        sg_layout.addWidget(btn_save_ass)
        sg_layout.addWidget(btn_save_copy)
        sp_layout.addWidget(save_group)
        
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.close)
        sp_layout.addWidget(btn_close)
        
        self.browser = QWebEngineView()
        if not WEB_ENGINE_AVAILABLE:
            self.browser.setOpenExternalLinks(False)
        
        self.content_layout.addWidget(self.settings_panel)
        self.content_layout.addWidget(self.browser)
        
    def on_page_loaded(self, ok):
        if ok:
            QTimer.singleShot(100, lambda: self.browser.page().runJavaScript("window.updateScrollStatus();"))
    
    def update_counter_label(self, index, total):
        self.current_h_index = index
        if total > 0:
            self.lbl_h_count.setText(f"{index + 1} / {total}")
        else:
            self.lbl_h_count.setText("0 / 0")

    def scroll_to_highlight(self, direction):
        js_get_total = "document.querySelectorAll('.highlighted-block').length"
        
        def navigate(total):
            if not total or int(total) == 0: return
            total = int(total)
            
            if direction == "next":
                target_index = self.current_h_index + 1
                if target_index >= total: target_index = 0
            else:
                target_index = self.current_h_index - 1
                if target_index < 0: target_index = total - 1

            js_jump = f"""
            (function() {{
                var blocks = document.querySelectorAll('.highlighted-block');
                var target = blocks[{target_index}];
                if (target) {{
                    blocks.forEach(b => b.classList.remove('active-replica'));
                    target.classList.add('active-replica');
                    target.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
                }}
            }})();
            """
            self.browser.page().runJavaScript(js_jump)
            self.current_h_index = target_index
            self.lbl_h_count.setText(f"{target_index + 1} / {total}")

        self.browser.page().runJavaScript(js_get_total, navigate)

    def update_preview(self):
        lines = self.main_app.get_episode_lines(self.ep_num)
        if not lines:
            self.browser.setHtml("<h3>Нет данных в серии</h3>")
            return
        
        try: self.browser.loadFinished.disconnect(self.on_page_loaded)
        except: pass
        self.browser.loadFinished.connect(self.on_page_loaded)

        cfg = self.main_app.data["export_config"]
        local_layout = self.combo_layout.currentText()
        processed = self.main_app.process_merge_logic(lines, cfg)
        html = self.main_app.generate_html_body(self.ep_num, processed, cfg, self.highlight_ids, override_layout=local_layout)
        self.browser.setHtml(html)
    
    def toggle_sidebar(self):
        is_hidden = self.settings_panel.isVisible()
        self.settings_panel.setVisible(not is_hidden)
        if is_hidden: self.btn_toggle_sidebar.setText("➡ Показать настройки")
        else: self.btn_toggle_sidebar.setText("⬅ Скрыть настройки")

    def on_setting_change(self):
        cfg = self.main_app.data["export_config"]
        cfg["f_time"] = self.s_time.value()
        cfg["f_char"] = self.s_char.value()
        cfg["f_actor"] = self.s_actor.value()
        cfg["f_text"] = self.s_text.value()
        # НЕ устанавливаем флаг грязности, так как это только изменение отображения, а не текста
        # self.main_app.set_dirty(True)  # ← Удалено
        self.update_preview()

    def open_actor_filter(self):
        all_aids = list(self.main_app.data["actors"].keys())
        current_selection = self.highlight_ids if self.highlight_ids is not None else all_aids
        d = ActorFilterDialog(self.main_app.data["actors"], current_selection, self)
        if d.exec():
            selected = d.get_selected()
            if len(selected) == len(all_aids) or len(selected) == 0:
                self.highlight_ids = None
            else:
                self.highlight_ids = selected
            self.update_preview()

    def save_to_original_ass(self):
        if QMessageBox.question(self, "Подтверждение",
            "Это перезапишет исходный файл .ass на диске.\nПродолжить?",
            QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            if self.main_app.save_episode_to_ass(self.ep_num):
                self._has_text_changes = False
                QMessageBox.information(self, "Успех", "Файл успешно сохранен!")

    def save_ass_copy(self):
        fn, _ = QFileDialog.getSaveFileName(self, "Сохранить копию", f"Episode_{self.ep_num}_edit.ass", "ASS Files (*.ass)")
        if fn:
            if self.main_app.save_episode_to_ass(self.ep_num, fn):
                QMessageBox.information(self, "Успех", f"Копия сохранена:\n{fn}")

    def closeEvent(self, event):
        # Проверяем только реальные изменения текста в этом окне предпросмотра,
        # а не глобальный флаг is_dirty который может быть установлен изменениями отображения
        if self._has_text_changes:
            reply = QMessageBox.question(self, "Несохраненные изменения",
                "У вас есть несохраненные изменения в тексте.\nХотите сохранить их в .ASS перед выходом?",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
            if reply == QMessageBox.Yes:
                if self.main_app.save_episode_to_ass(self.ep_num):
                    self._has_text_changes = False
                    event.accept()
                else: 
                    event.ignore()
            elif reply == QMessageBox.No: 
                event.accept()
            else: 
                event.ignore()
        else: 
            event.accept()

# --- ГЛОБАЛЬНЫЙ ПОИСК ---
class GlobalSearchDialog(QDialog):
    def __init__(self, project_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Глобальный поиск по проекту")
        self.resize(900, 600)
        self.project_data = project_data
        self.main_app = parent
        layout = QVBoxLayout(self)
        
        search_panel = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите текст или имя персонажа...")
        self.search_input.returnPressed.connect(self.perform_search)
        btn_search = QPushButton("Найти")
        btn_search.clicked.connect(self.perform_search)
        search_panel.addWidget(QLabel("Поиск:"))
        search_panel.addWidget(self.search_input)
        search_panel.addWidget(btn_search)
        layout.addLayout(search_panel)
        
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Серия", "Таймкод", "Персонаж", "Текст"])
        customize_table(self.table)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.cellDoubleClicked.connect(self.go_to_result)
        layout.addWidget(self.table)

    def perform_search(self):
        query = self.search_input.text().lower().strip()
        if not query: return
        self.table.setRowCount(0)
        episodes = self.project_data.get("episodes", {})
        for ep_num in sorted(episodes.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            path = episodes[ep_num]
            if not os.path.exists(path): continue
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.startswith("Dialogue:"):
                            parts = line.split(',', 9)
                            if len(parts) > 9:
                                start_time = parts[1]
                                char_name = parts[4].strip()
                                text_clean = re.sub(r'\{.*?\}', '', parts[9]).strip()
                                if query in char_name.lower() or query in text_clean.lower():
                                    self.add_result_row(ep_num, start_time, char_name, text_clean)
            except: pass
        if self.table.rowCount() == 0: QMessageBox.information(self, "Поиск", "Ничего не найдено.")

    def add_result_row(self, ep, time, char, text):
        row = self.table.rowCount()
        self.table.insertRow(row)
        item_ep = QTableWidgetItem(str(ep)); item_ep.setData(Qt.UserRole, ep)
        self.table.setItem(row, 0, item_ep)
        self.table.setItem(row, 1, QTableWidgetItem(time))
        self.table.setItem(row, 2, QTableWidgetItem(char))
        self.table.setItem(row, 3, QTableWidgetItem(text))

    def go_to_result(self, row, col):
        ep_num = self.table.item(row, 0).data(Qt.UserRole)
        if self.main_app: self.main_app.switch_to_episode(ep_num)

# --- ЦВЕТА ---
class CustomColorDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выберите цвет")
        self.selected_color = None
        layout = QVBoxLayout(self)
        grid = QGridLayout()
        r, c = 0, 0
        for color_hex in MY_PALETTE:
            btn = QPushButton()
            btn.setFixedSize(35, 35)
            btn.setStyleSheet(f"background-color: {color_hex}; border-radius: 4px; border: 1px solid #999;")
            btn.clicked.connect(lambda ch=False, clr=color_hex: self.select_color(clr))
            grid.addWidget(btn, r, c)
            c += 1
            if c > 4: c = 0; r += 1
        layout.addLayout(grid)
        btn_custom = QPushButton("Другой цвет...")
        btn_custom.clicked.connect(self.open_system_picker)
        layout.addWidget(btn_custom)

    def select_color(self, clr):
        self.selected_color = clr
        self.accept()

    def open_system_picker(self):
        c = QColorDialog.getColor()
        if c.isValid(): self.selected_color = c.name(); self.accept()

# --- НАСТРОЙКИ ЭКСПОРТА ---
class ExportSettingsDialog(QDialog):
    def __init__(self, current_settings, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки монтажного листа")
        self.settings = current_settings
        self.parent_app = parent
        self.highlight_ids_export = self.settings.get('highlight_ids_export')
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.layout_type = QComboBox(); self.layout_type.addItems(["Таблица", "Сценарий"])
        self.layout_type.setCurrentText(self.settings.get('layout_type', "Таблица"))
        form.addRow("Тип разметки:", self.layout_type)
        layout.addLayout(form)
        
        col_group = QGroupBox("Элементы в листе")
        c_lay = QVBoxLayout(col_group)
        self.col_tc = QCheckBox("Таймкоды"); self.col_tc.setChecked(self.settings.get('col_tc', True))
        self.col_char = QCheckBox("Имя персонажа"); self.col_char.setChecked(self.settings.get('col_char', True))
        self.col_actor = QCheckBox("Имя актера"); self.col_actor.setChecked(self.settings.get('col_actor', True))
        self.col_text = QCheckBox("Текст реплики"); self.col_text.setChecked(self.settings.get('col_text', True))
        for cb in [self.col_tc, self.col_char, self.col_actor, self.col_text]: c_lay.addWidget(cb)
        layout.addWidget(col_group)

        font_group = QGroupBox("Размеры шрифтов (px)")
        f_lay = QFormLayout(font_group)
        self.f_time = QSpinBox(); self.f_time.setValue(self.settings.get('f_time', 12))
        self.f_char = QSpinBox(); self.f_char.setValue(self.settings.get('f_char', 14))
        self.f_actor = QSpinBox(); self.f_actor.setValue(self.settings.get('f_actor', 14))
        self.f_text = QSpinBox(); self.f_text.setValue(self.settings.get('f_text', 16))
        f_lay.addRow("Таймкод:", self.f_time); f_lay.addRow("Персонаж:", self.f_char)
        f_lay.addRow("Актер:", self.f_actor); f_lay.addRow("Текст:", self.f_text)
        layout.addWidget(font_group)
        
        l_lay = QFormLayout()
        self.merge_gap = QSpinBox(); self.merge_gap.setValue(self.settings.get('merge_gap', 5))
        self.p_short = QDoubleSpinBox(); self.p_short.setValue(self.settings.get('p_short', 0.5))
        self.p_long = QDoubleSpinBox(); self.p_long.setValue(self.settings.get('p_long', 2.0))
        l_lay.addRow("Порог слияния (сек):", self.merge_gap)
        l_lay.addRow("Пауза для '/' (сек):", self.p_short)
        l_lay.addRow("Пауза для '//' (сек):", self.p_long)
        layout.addLayout(l_lay)
        
        # Группа для выбора цветов и актёров
        color_group = QGroupBox("Отображение цветов")
        color_lay = QVBoxLayout(color_group)
        
        self.use_color = QCheckBox("Использовать цвета актёров"); self.use_color.setChecked(self.settings.get('use_color', True))
        color_lay.addWidget(self.use_color)
        
        # Кнопка для выбора актёров
        btn_filter = QPushButton("Выбрать актёров для подсветки...")
        btn_filter.clicked.connect(self.open_actor_filter_for_export)
        color_lay.addWidget(btn_filter)
        
        layout.addWidget(color_group)
        
        self.round_time = QCheckBox("Округлять время"); self.round_time.setChecked(self.settings.get('round_time', False))
        self.open_auto = QCheckBox("Открыть после экспорта"); self.open_auto.setChecked(self.settings.get('open_auto', True))
        for cb in [self.round_time, self.open_auto]: layout.addWidget(cb)
        
        b_ok = QPushButton("Сохранить настройки"); b_ok.clicked.connect(self.accept)
        layout.addWidget(b_ok)

    def open_actor_filter_for_export(self):
        """Открытие диалога выбора актёров для экспорта"""
        if not self.parent_app or not hasattr(self.parent_app, 'data'):
            QMessageBox.warning(self, "Ошибка", "Не удалось получить доступ к данным актёров")
            return
        
        all_aids = list(self.parent_app.data["actors"].keys())
        current_selection = self.highlight_ids_export if self.highlight_ids_export is not None else all_aids
        d = ActorFilterDialog(self.parent_app.data["actors"], current_selection, self)
        if d.exec():
            selected = d.get_selected()
            if len(selected) == len(all_aids) or len(selected) == 0:
                self.highlight_ids_export = None
            else:
                self.highlight_ids_export = selected

    def get_settings(self):
        s = self.settings.copy()
        s.update({
            'layout_type': self.layout_type.currentText(),
            'col_tc': self.col_tc.isChecked(), 'col_char': self.col_char.isChecked(),
            'col_actor': self.col_actor.isChecked(), 'col_text': self.col_text.isChecked(),
            'f_time': self.f_time.value(), 'f_char': self.f_char.value(),
            'f_actor': self.f_actor.value(), 'f_text': self.f_text.value(),
            'merge_gap': self.merge_gap.value(), 'p_short': self.p_short.value(),
            'p_long': self.p_long.value(), 'use_color': self.use_color.isChecked(),
            'round_time': self.round_time.isChecked(), 'open_auto': self.open_auto.isChecked(),
            'highlight_ids_export': self.highlight_ids_export
        })
        return s

# --- ОТЧЕТ ---
class SummaryDialog(QDialog):
    def __init__(self, data, target_ep=None, parent=None):
        super().__init__(parent)
        self.target_ep = target_ep
        self.setWindowTitle(f"Отчет: {'Серия ' + target_ep if target_ep else 'Проект'}")
        self.resize(1000, 700)
        layout = QVBoxLayout(self)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Актер", "Цвет", "Колец", "Слов", "Персонажи"])
        customize_table(self.table)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        layout.addWidget(self.table)
        self.calculate_stats(data)
        b_close = QPushButton("Закрыть"); b_close.clicked.connect(self.accept)
        layout.addWidget(b_close)

    def calculate_stats(self, data):
        gap = data["export_config"].get('merge_gap', 5)
        stats = {aid: {"rings": 0, "words": 0, "roles": set()} for aid in data["actors"]}
        unassigned = {"rings": 0, "words": 0, "roles": set()}
        eps = {self.target_ep: data["episodes"][self.target_ep]} if self.target_ep else data["episodes"]
        for ep_num, path in eps.items():
            if not os.path.exists(path): continue
            lines = []
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.startswith("Dialogue:"):
                            p = line.split(',', 9); t = re.sub(r'\{.*?\}', '', p[9]).strip()
                            if t: lines.append({'s': ass_time_to_seconds(p[1]), 'e': ass_time_to_seconds(p[2]), 'char': p[4].strip(), 'text': t})
            except: continue
            if not lines: continue
            merged = []; curr = lines[0]
            for i in range(1, len(lines)):
                nxt = lines[i]
                if nxt['char'] == curr['char'] and (nxt['s'] - curr['e']) < gap:
                    curr['text'] += " " + nxt['text']; curr['e'] = nxt['e']
                else: merged.append(curr); curr = nxt
            merged.append(curr)
            for l in merged:
                aid = data["global_map"].get(l['char'])
                target = stats[aid] if aid in stats else unassigned
                target["rings"] += 1; target["words"] += len(l['text'].split()); target["roles"].add(l['char'])
        for aid, s in stats.items():
            if s["rings"] == 0 and self.target_ep: continue
            row = self.table.rowCount(); self.table.insertRow(row)
            info = data["actors"][aid]
            self.table.setItem(row, 0, QTableWidgetItem(info["name"]))
            c_it = QTableWidgetItem(); c_it.setBackground(QColor(info["color"])); self.table.setItem(row, 1, c_it)
            self.table.setItem(row, 2, QTableWidgetItem(str(s["rings"])))
            self.table.setItem(row, 3, QTableWidgetItem(str(s["words"])))
            roles_item = QTableWidgetItem("\n".join([f"• {r}" for r in sorted(list(s["roles"]))]))
            self.table.setItem(row, 4, roles_item)
        if unassigned["roles"]:
            row = self.table.rowCount(); self.table.insertRow(row)
            item_name = QTableWidgetItem("НЕ РАСПРЕДЕЛЕНЫ"); item_name.setForeground(QColor("red"))
            self.table.setItem(row, 0, item_name)
            self.table.setItem(row, 2, QTableWidgetItem(str(unassigned["rings"])))
            self.table.setItem(row, 3, QTableWidgetItem(str(unassigned["words"])))
            self.table.setItem(row, 4, QTableWidgetItem(", ".join(sorted(list(unassigned["roles"])))))

# --- ВИДЕО ---
class VideoPreviewWindow(QDialog):
    def __init__(self, video_path, lines, ep_num, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Просмотр: Серия {ep_num}")
        self.resize(1000, 800)
        self.video_path = video_path; self.lines = lines; self.init_ui()
        if self.video_path and os.path.exists(self.video_path):
            self.media_player.setSource(QUrl.fromLocalFile(os.path.abspath(self.video_path)))

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.video_widget = QVideoWidget(); self.video_widget.setMinimumHeight(400)
        layout.addWidget(self.video_widget)
        self.media_player = QMediaPlayer(); self.audio_output = QAudioOutput()
        self.media_player.setVideoOutput(self.video_widget); self.media_player.setAudioOutput(self.audio_output)
        ctrl = QHBoxLayout()
        btn_p = QPushButton("Play/Pause"); btn_p.clicked.connect(self.toggle_play)
        self.slider = QSlider(Qt.Horizontal)
        self.media_player.positionChanged.connect(self.slider.setValue)
        self.media_player.durationChanged.connect(lambda d: self.slider.setRange(0, d))
        self.slider.sliderMoved.connect(self.media_player.setPosition)
        ctrl.addWidget(btn_p); ctrl.addWidget(self.slider); layout.addLayout(ctrl)
        self.line_table = QTableWidget(0, 3)
        self.line_table.setHorizontalHeaderLabels(["Время", "Персонаж", "Текст"])
        customize_table(self.line_table); self.line_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.line_table.cellClicked.connect(self.seek_to_line)
        layout.addWidget(self.line_table)
        for line in self.lines:
            row = self.line_table.rowCount(); self.line_table.insertRow(row)
            self.line_table.setItem(row, 0, QTableWidgetItem(format_seconds_to_tc(line['s'])))
            self.line_table.setItem(row, 1, QTableWidgetItem(line['char']))
            self.line_table.setItem(row, 2, QTableWidgetItem(line['text']))
            self.line_table.item(row, 0).setData(Qt.UserRole, line['s'])

    def toggle_play(self):
        if self.media_player.playbackState() == QMediaPlayer.PlayingState: self.media_player.pause()
        else: self.media_player.play()
    def seek_to_line(self, row, col):
        pos = self.line_table.item(row, 0).data(Qt.UserRole)
        self.media_player.setPosition(int(pos * 1000)); self.media_player.play()
    def closeEvent(self, e): self.media_player.stop(); e.accept()

# --- РОЛИ ---
class ActorRolesDialog(QDialog):
    def __init__(self, actor_name, current_roles, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Роли: {actor_name}")
        layout = QVBoxLayout(self)
        self.list_widget = QListWidget(); self.list_widget.addItems(current_roles)
        layout.addWidget(self.list_widget)
        btns = QHBoxLayout(); b_add = QPushButton("Добавить"); b_del = QPushButton("Удалить")
        btns.addWidget(b_add); btns.addWidget(b_del); layout.addLayout(btns)
        b_add.clicked.connect(self.add_role); b_del.clicked.connect(self.del_role)
        b_ok = QPushButton("Готово"); b_ok.clicked.connect(self.accept); layout.addWidget(b_ok)
    def add_role(self):
        r, ok = QInputDialog.getText(self, "Новая роль", "Имя:")
        if ok and r: self.list_widget.addItem(r)
    def del_role(self):
        for i in self.list_widget.selectedItems(): self.list_widget.takeItem(self.list_widget.row(i))
    def get_roles(self): return [self.list_widget.item(i).text() for i in range(self.list_widget.count())]

# --- MAIN APP ---
class DubbingApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Dubbing Manager")
        self.resize(1350, 850)
        self.setAcceptDrops(True)
        self.current_project_path = None
        self.is_dirty = False
        self.sort_col = 1
        self.sort_desc = True
        self.preview_window = None
        self.teleprompter_window = None 
        
        self.data = {
            "project_name": "Новый проект", 
            "actors": {}, "global_map": {}, "episodes": {}, "video_paths": {}, 
            "export_config": {
                'layout_type': 'Таблица', 'col_tc': True, 'col_char': True, 
                'col_actor': True, 'col_text': True, 'f_time': 21, 'f_char': 20, 
                'f_actor': 14, 'f_text': 30, 'use_color': True, 'merge': True, 
                'merge_gap': 5, 'p_short': 0.5, 'p_long': 2.0, 
                'open_auto': True, 'round_time': False
            },
            # НОВЫЙ БЛОК: Настройки суфлёра
            "prompter_config": {
            "f_tc": 20, "f_char": 24, "f_actor": 18, "f_text": 36,
            "focus_ratio": 0.5, "is_mirrored": False, "show_header": False,
            "port_in": 8000, "port_out": 9000, "sync_in": True, "sync_out": False,
            "key_prev": "Left",  # Горячая клавиша назад
            "key_next": "Right", # Горячая клавиша вперед
            "colors": {
                "bg": "#000000", "active_text": "#FFFFFF", "inactive_text": "#444444",
                "tc": "#888888", "actor": "#AAAAAA", "header_bg": "#111111", "header_text": "#00FF00"
            }
        }
        }
        self.current_ep_stats = []
        self.character_names_changed = {}  # Отслеживание: ep_num -> True/False
        self.init_ui()
        self.update_window_title()
        self.autosave_timer = QTimer(self); self.autosave_timer.timeout.connect(self.auto_save); self.autosave_timer.start(300000)

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        
        # --- ЛЕВАЯ ПАНЕЛЬ (СПИСОК АКТЕРОВ) ---
        left_panel = QVBoxLayout()
        left_widget = QFrame()
        left_widget.setFixedWidth(350)
        left_widget.setFrameShape(QFrame.StyledPanel)
        left_widget.setLayout(left_panel)
        
        left_panel.addWidget(QLabel("<b>БАЗА АКТЕРОВ</b>"))
        self.actor_table = QTableWidget(0, 3)
        self.actor_table.setHorizontalHeaderLabels(["Актер", "Цвет", "Роли"])
        customize_table(self.actor_table)
        self.actor_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.actor_table.itemChanged.connect(self.on_actor_renamed)
        self.actor_table.cellClicked.connect(self.on_actor_cell_clicked)
        left_panel.addWidget(self.actor_table)
        
        b_add = QPushButton("+ Актер")
        b_add.clicked.connect(self.add_actor_dialog)
        left_panel.addWidget(b_add)
        
        b_sum = QPushButton("📋 Сводный отчет проекта")
        b_sum.clicked.connect(self.show_project_summary)
        left_panel.addWidget(b_sum)
        
        main_layout.addWidget(left_widget)

        # --- ПРАВАЯ (ЦЕНТРАЛЬНАЯ) ЧАСТЬ ---
        right_panel = QVBoxLayout()
        
        # 1. ВЕРХНЯЯ СТРОКА (ПРОЕКТ)
        top = QHBoxLayout()
        self.proj_edit = QLineEdit()
        self.proj_edit.textChanged.connect(self.on_project_name_changed)
        top.addWidget(QLabel("Проект:"))
        top.addWidget(self.proj_edit)
        
        b_load = QPushButton("Открыть")
        b_load.clicked.connect(self.load_project_dialog)
        top.addWidget(b_load)
        
        b_save = QPushButton("Сохранить")
        b_save.clicked.connect(self.save_project)
        top.addWidget(b_save)
        
        b_copy = QPushButton("Копия")
        b_copy.clicked.connect(self.save_project_as)
        top.addWidget(b_copy)
        
        right_panel.addLayout(top)
        
        # 2. УПРАВЛЕНИЕ СЕРИЯМИ И ПОИСК
        ep_ctrl = QHBoxLayout()
        
        self.ep_combo = QComboBox()
        self.ep_combo.setMinimumWidth(120)
        self.ep_combo.currentIndexChanged.connect(self.change_episode)
        ep_ctrl.addWidget(QLabel("Серия:"))
        ep_ctrl.addWidget(self.ep_combo)
        
        b_ren = QPushButton("✎")
        b_ren.setFixedWidth(30)
        b_ren.clicked.connect(self.rename_episode)
        ep_ctrl.addWidget(b_ren)
        
        b_ass = QPushButton("+ .ASS")
        b_ass.clicked.connect(lambda: self.import_ass())
        ep_ctrl.addWidget(b_ass)
        
        self.btn_save_ass = QPushButton()
        self.btn_save_ass.setFixedWidth(120)
        self.btn_save_ass.clicked.connect(self.save_current_episode_ass)
        self.update_save_ass_button()
        ep_ctrl.addWidget(self.btn_save_ass)
        
        b_vid = QPushButton("🎬 Видео")
        b_vid.clicked.connect(self.set_episode_video)
        ep_ctrl.addWidget(b_vid)
        
        b_ep_sum = QPushButton("📊 Отчет серии")
        b_ep_sum.clicked.connect(self.show_episode_summary)
        ep_ctrl.addWidget(b_ep_sum)
        
        # Распорка, чтобы отодвинуть поиск вправо
        ep_ctrl.addStretch()
        
        # Поле поиска (стандартный виджет без стилей)
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Поиск...")
        self.search_edit.setFixedWidth(160) # Немного шире для удобства
        self.search_edit.textChanged.connect(self.refresh_main_table)
        ep_ctrl.addWidget(self.search_edit)
        
        b_glob_search = QPushButton("🔍 Глобальный поиск")
        b_glob_search.clicked.connect(self.open_global_search)
        ep_ctrl.addWidget(b_glob_search)
        
        self.filter_unassigned = QCheckBox("Пустые")
        self.filter_unassigned.toggled.connect(self.refresh_main_table)
        ep_ctrl.addWidget(self.filter_unassigned)
        
        right_panel.addLayout(ep_ctrl)
        
        # 3. ЦЕНТРАЛЬНАЯ ОБЛАСТЬ (ТАБЛИЦА + БОКОВАЯ ПАНЕЛЬ ИНСТРУМЕНТОВ)
        middle_layout = QHBoxLayout()
        
        # Стек таблиц
        self.table_stack = QStackedWidget()
        self.main_table = QTableWidget(0, 6)
        self.main_table.setHorizontalHeaderLabels(["Персонаж", "Строчек", "Колец", "Слов", "Актер", "📺"])
        customize_table(self.main_table)
        self.main_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.main_table.horizontalHeader().setSectionsClickable(True)
        self.main_table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.main_table.itemChanged.connect(self.on_character_name_changed)
        
        self.missing_file_widget = QWidget()
        mf_lay = QVBoxLayout(self.missing_file_widget)
        self.lbl_missing = QLabel("ФАЙЛ НЕ НАЙДЕН")
        self.lbl_missing.setStyleSheet("color: red; font-weight: bold;") # Здесь стиль оставляем для акцента на ошибке
        self.lbl_missing.setAlignment(Qt.AlignCenter)
        b_relink = QPushButton("Найти...")
        b_relink.clicked.connect(self.relink_file)
        mf_lay.addStretch()
        mf_lay.addWidget(self.lbl_missing)
        mf_lay.addWidget(b_relink)
        mf_lay.addStretch()
        
        self.table_stack.addWidget(self.main_table)
        self.table_stack.addWidget(self.missing_file_widget)
        
        middle_layout.addWidget(self.table_stack, stretch=1)
        
        # --- БОКОВАЯ ПАНЕЛЬ ИНСТРУМЕНТОВ ---
        tools_sidebar_widget = QWidget()
        tools_sidebar_widget.setFixedWidth(160)
        tools_sidebar_layout = QVBoxLayout(tools_sidebar_widget)
        tools_sidebar_layout.setContentsMargins(5, 0, 0, 0)
        
        tools_sidebar_layout.addWidget(QLabel("<b>Инструменты:</b>"))
        
        b_all_v = QPushButton("📺 Просмотр серии")
        b_all_v.clicked.connect(lambda: self.open_preview(None))
        tools_sidebar_layout.addWidget(b_all_v)
        
        b_live_html = QPushButton("📃 Монтажный лист")
        b_live_html.clicked.connect(self.open_live_preview)
        tools_sidebar_layout.addWidget(b_live_html)
        
        b_prompter = QPushButton("🎤 Телесуфлёр")
        b_prompter.clicked.connect(self.open_teleprompter)
        tools_sidebar_layout.addWidget(b_prompter)
        
        b_reaper = QPushButton("🎹 Reaper RPP")
        b_reaper.clicked.connect(self.export_to_reaper_rpp)
        tools_sidebar_layout.addWidget(b_reaper)
        
        tools_sidebar_layout.addStretch()
        middle_layout.addWidget(tools_sidebar_widget)
        
        right_panel.addLayout(middle_layout)
        
        # 4. НИЖНЯЯ ПАНЕЛЬ
        bottom_panel = QHBoxLayout()
        
        # Слева: Кнопка назначения
        b_bulk = QPushButton("⚡ Назначить выделенным")
        b_bulk.clicked.connect(self.bulk_assign_actor)
        bottom_panel.addWidget(b_bulk)
        
        # Распорка по центру
        bottom_panel.addStretch()
        
        # Справа: Настройки, затем Экспорт
        b_cfg = QPushButton("⚙ Настройки")
        b_cfg.clicked.connect(self.open_export_settings)
        bottom_panel.addWidget(b_cfg)
        
        # Группа экспорта
        exp_group = QGroupBox("Экспорт")
        exp_lay = QHBoxLayout(exp_group)
        exp_lay.setContentsMargins(5, 5, 5, 5) # Немного компактнее
        
        self.chk_exp_html = QCheckBox("Лист")
        self.chk_exp_html.setChecked(True)
        self.chk_exp_xls = QCheckBox("Excel")
        
        self.radio_cur = QRadioButton("Текущая")
        self.radio_cur.setChecked(True)
        self.radio_all = QRadioButton("Все")
        
        self.btn_run_export = QPushButton("ЭКСПОРТ")
        self.btn_run_export.clicked.connect(self.run_unified_export)
        
        exp_lay.addWidget(self.chk_exp_html)
        exp_lay.addWidget(self.chk_exp_xls)
        exp_lay.addSpacing(10)
        exp_lay.addWidget(self.radio_cur)
        exp_lay.addWidget(self.radio_all)
        exp_lay.addSpacing(10)
        exp_lay.addWidget(self.btn_run_export)
        
        bottom_panel.addWidget(exp_group)
        
        right_panel.addLayout(bottom_panel)
        main_layout.addLayout(right_panel)
        
    def hex_to_reaper_color(self, hex_color):
        """Конвертация HEX в BGR Int для Reaper"""
        if not hex_color or not hex_color.startswith('#'):
            return 0
        c = QColor(hex_color)
        if not c.isValid():
            return 0
        # Бит 0x01000000 (24-й бит) включает Custom Color в Reaper
        # Порядок цветов BGR
        val = 0x01000000 | (c.blue() << 16) | (c.green() << 8) | c.red()
        return val
        
    def export_to_reaper_rpp(self):
        """
        Финальная версия экспорта в Reaper (.RPP).
        Гарантирует правильное создание РЕГИОНОВ с текстом.
        """
        ep_num = self.ep_combo.currentData()
        if not ep_num:
            QMessageBox.warning(self, "Ошибка", "Выберите серию.")
            return

        # 1. Видео и диалог
        video_path = self.data["video_paths"].get(ep_num)
        
        # Убедитесь, что класс ReaperExportDialog добавлен в код (см. предыдущий ответ)
        dlg = ReaperExportDialog(video_path, self)
        if dlg.exec() != QDialog.Accepted:
            return

        use_video, use_regions = dlg.get_options()

        if use_video and video_path:
            video_path = os.path.abspath(video_path)

        # 2. Путь сохранения
        default_name = f"{self.data.get('project_name', 'Project')} - Ep{ep_num}.rpp"
        save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить RPP", default_name, "Reaper Project (*.rpp)")
        if not save_path: return

        # 3. Данные
        lines = self.get_episode_lines(ep_num)
        active_actor_ids = set()
        
        max_time = 600.0
        if lines:
            max_time = max(l['e'] for l in lines) + 600.0
            for line in lines:
                aid = self.data["global_map"].get(line['char'])
                if aid: active_actor_ids.add(aid)

        # Склейка реплик
        processed_lines = self.process_merge_logic(lines, self.data["export_config"])

        # --- ЗАПИСЬ RPP ---
        rpp = []
        # Заголовок проекта
        rpp.append('<REAPER_PROJECT 0.1 "7.0"')

        # --- БЛОК РЕГИОНОВ ---
        if use_regions:
            for i, line in enumerate(processed_lines):
                start = float(line['s'])
                end = float(line['e'])
                
                # ЗАЩИТА ОТ НУЛЕВОЙ ДЛИНЫ
                # Если регион короче 0.5 сек, делаем его минимум 2 секунды, чтобы было видно
                if (end - start) < 0.5:
                    end = start + 2.0

                char = line['char']
                # Очистка текста от кавычек (ломают RPP) и переносов
                safe_text = line['text'].replace('"', "'").replace('\n', ' ').strip()
                label = f"{char}: {safe_text}"
                
                # Цвет
                aid = self.data["global_map"].get(char)
                color_int = 0
                if aid and aid in self.data["actors"]:
                    # Используем int() для гарантии
                    color_int = int(self.hex_to_reaper_color(self.data["actors"][aid]["color"]))

                # ФОРМАТ СТРОКИ REAPER:
                # MARKER <ID> <Start> "Name" <IsRegion> <Color> <End>
                # 1 (IsRegion) - делает это регионом
                # {end:.4f} - это ВРЕМЯ ОКОНЧАНИЯ, а не длина
                rpp.append(f'  MARKER {i+1} {start:.4f} "{label}" 1 {color_int} {end:.4f}')

        # --- ВИДЕО ---
        if use_video and video_path:
            rpp.append('  <TRACK')
            rpp.append('    NAME "VIDEO"')
            rpp.append('    <ITEM')
            rpp.append('      POSITION 0.0')
            rpp.append('      LOOP 0')
            rpp.append(f'      LENGTH {max_time:.4f}')
            rpp.append('      <SOURCE VIDEO')
            rpp.append(f'        FILE "{video_path}"')
            rpp.append('      >')
            rpp.append('    >')
            rpp.append('  >')

        # --- ДОРОЖКИ АКТЕРОВ ---
        sorted_actors = []
        for aid in active_actor_ids:
            if aid in self.data["actors"]:
                sorted_actors.append(self.data["actors"][aid])
        sorted_actors.sort(key=lambda x: x['name'])

        for actor in sorted_actors:
            color_int = int(self.hex_to_reaper_color(actor['color']))
            rpp.append('  <TRACK')
            rpp.append(f'    NAME "{actor["name"]}"')
            rpp.append(f'    PEAKCOL {color_int}')
            rpp.append('    REC 0')
            rpp.append('    SHOWINMIX 1')
            rpp.append('  >')

        rpp.append('>') # End Project

        # --- СОХРАНЕНИЕ ---
        try:
            with open(save_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(rpp))
            
            if QMessageBox.question(self, "Готово", "Проект создан. Открыть в Reaper?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                if sys.platform == 'win32': os.startfile(save_path)
                elif sys.platform == 'darwin': os.system(f'open "{save_path}"')
                else: os.system(f'xdg-open "{save_path}"')
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить: {e}")

    def switch_to_episode(self, ep_num):
        index = self.ep_combo.findData(ep_num)
        if index >= 0: self.ep_combo.setCurrentIndex(index)

    def open_global_search(self):
        GlobalSearchDialog(self.data, self).exec()
        
    def open_live_preview(self):
        ep = self.ep_combo.currentData()
        if not ep: QMessageBox.information(self, "Инфо", "Выберите серию."); return
        if self.preview_window is not None: self.preview_window.close()
        self.preview_window = HtmlLivePreview(self, ep); self.preview_window.show()

    def open_teleprompter(self):
        ep = self.ep_combo.currentData()
        if not ep: QMessageBox.information(self, "Инфо", "Выберите серию."); return
        if self.teleprompter_window is not None: self.teleprompter_window.close()
        self.teleprompter_window = TeleprompterWindow(self, ep); self.teleprompter_window.show()

    def set_dirty(self, d=True): self.is_dirty = d; self.update_window_title()
    def update_window_title(self):
        t = "Dubbing Manager"
        if self.current_project_path: t += f" - {os.path.basename(self.current_project_path)}"
        else: t += " - [Новый]"
        if self.is_dirty: t += "*"
        self.setWindowTitle(t)

    def maybe_save(self):
        if not self.is_dirty: return True
        r = QMessageBox.question(self, "Сохранить?", "Сохранить изменения?", QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
        if r == QMessageBox.Save: return self.save_project()
        return r == QMessageBox.Discard

    def closeEvent(self, e): 
        if self.maybe_save(): e.accept()
        else: e.ignore()

    def auto_save(self):
        if not self.is_dirty: return
        p = (self.current_project_path + ".bak") if self.current_project_path else "temp_autosave.json.bak"
        with open(p, 'w', encoding='utf-8') as f: json.dump(self.data, f, ensure_ascii=False, indent=4)

    def on_project_name_changed(self, t): self.data["project_name"] = t; self.set_dirty()

    def on_character_name_changed(self, item):
        """Обработчик изменения имени персонажа в таблице"""
        if item.column() != 0:  # Только первая колонка (Персонаж)
            return
        
        ep = self.ep_combo.currentData()
        if not ep:
            return
        
        old_name = item.data(Qt.UserRole)  # Оригинальное имя
        new_name = item.text().strip()
        
        if new_name == old_name or not new_name:
            return  # Никаких изменений или пустое имя
        
        # Убедимся что loaded_episodes загружена
        self.get_episode_lines(ep)
        
        # Обновляем данные в памяти
        # Если персонаж был заказан актёру, обновляем маппинг
        if old_name in self.data["global_map"]:
            aid = self.data["global_map"][old_name]
            del self.data["global_map"][old_name]
            self.data["global_map"][new_name] = aid
        
        # Обновляем loaded_episodes если они есть
        if ep in self.data.get("loaded_episodes", {}):
            for line in self.data["loaded_episodes"][ep]:
                if line['char'] == old_name:
                    line['char'] = new_name
        
        # Обновляем текущие статистики
        for stat in self.current_ep_stats:
            if stat["name"] == old_name:
                stat["name"] = new_name
                break
        
        # Обновляем UserRole для корректной работы при повторном редактировании
        item.setData(Qt.UserRole, new_name)
        
        # Отмечаем что имена персонажей изменились
        self.character_names_changed[ep] = True
        self.update_save_ass_button()
        
        # Обновляем другие таблицы и интерфейс
        self.refresh_actor_list()
        self.set_dirty(True)

    def update_save_ass_button(self):
        """Обновляет текст кнопки сохранения ASS с индикатором изменений"""
        ep = self.ep_combo.currentData()
        has_changes = self.character_names_changed.get(ep, False)
        
        if has_changes:
            self.btn_save_ass.setText("💾 Сохр.* ASS")
            self.btn_save_ass.setStyleSheet("font-weight: bold; color: red;")
        else:
            self.btn_save_ass.setText("💾 Сохранить")
            self.btn_save_ass.setStyleSheet("")

    def save_current_episode_ass(self):
        """Сохраняет текущую серию в ASS файл"""
        ep = self.ep_combo.currentData()
        if not ep:
            QMessageBox.warning(self, "Ошибка", "Выберите серию.")
            return
        
        if self.save_episode_to_ass(ep):
            # Очищаем флаг измеменений для этой серии
            self.character_names_changed[ep] = False
            self.update_save_ass_button()
            QMessageBox.information(self, "Успех", f"Серия {ep} сохранена в ASS файл.")

    def save_project(self): return self._do_save(self.current_project_path) if self.current_project_path else self.save_project_as()
    
    def save_project_as(self):
        p, _ = QFileDialog.getSaveFileName(self, "Сохранить", "", "*.json")
        if p: self.current_project_path = p; res = self._do_save(p); self.update_window_title(); return res
        return False

    def _do_save(self, p):
        try:
            with open(p, 'w', encoding='utf-8') as f: json.dump(self.data, f, ensure_ascii=False, indent=4)
            self.set_dirty(False); return True
        except: return False

    def load_project_dialog(self):
        if self.maybe_save():
            p, _ = QFileDialog.getOpenFileName(self, "Открыть", "", "*.json")
            if p: self._load_from_path(p)

    def _load_from_path(self, p):
        with open(p, 'r', encoding='utf-8') as f: self.data = json.load(f)
        if "video_paths" not in self.data: self.data["video_paths"] = {}
        if "export_config" not in self.data: self.data["export_config"] = {}
        self.current_project_path = p; self.proj_edit.setText(self.data.get("project_name", "Проект"))
        self.refresh_actor_list(); self.update_ep_list(); self.set_dirty(False)

    def on_header_clicked(self, i):
        if i > 3: return
        if self.sort_col == i: self.sort_desc = not self.sort_desc
        else: self.sort_col = i; self.sort_desc = True
        self.refresh_main_table()

    def add_actor_dialog(self):
        n, ok = QInputDialog.getText(self, "Новый актер", "Имя:")
        if ok and n:
            d = CustomColorDialog(self)
            if d.exec():
                self.data["actors"][str(datetime.now().timestamp())] = {"name":n, "color": d.selected_color or "#ffffff"}
                self.refresh_actor_list(); self.refresh_main_table(); self.set_dirty()

    def on_actor_cell_clicked(self, r, c):
        if c == 1:
            aid = self.actor_table.item(r, 0).data(Qt.UserRole)
            d = CustomColorDialog(self)
            if d.exec():
                if d.selected_color:
                    self.data["actors"][aid]["color"] = d.selected_color
                    self.refresh_actor_list(); self.refresh_main_table(); self.set_dirty()

    def on_actor_renamed(self, it):
        aid = it.data(Qt.UserRole)
        if aid: self.data["actors"][aid]["name"] = it.text(); self.refresh_main_table(); self.set_dirty()

    def bulk_assign_actor(self):
        sel = self.main_table.selectionModel().selectedRows()
        if not sel: return
        names = ["- Удалить -"] + [a["name"] for a in self.data["actors"].values()]; ids = [None] + list(self.data["actors"].keys())
        n, ok = QInputDialog.getItem(self, "Назначить", "Актер:", names, 0, False)
        if ok:
            aid = ids[names.index(n)]
            for idx in sel:
                c = self.main_table.item(idx.row(), 0).text()
                if aid: self.data["global_map"][c] = aid
                elif c in self.data["global_map"]: del self.data["global_map"][c]
            self.refresh_actor_list(); self.refresh_main_table(); self.set_dirty()

    def set_episode_video(self):
        ep = self.ep_combo.currentData()
        if ep:
            p, _ = QFileDialog.getOpenFileName(self, "Видео", "", "Video (*.mp4 *.mkv *.avi *.mov)")
            if p: self.data.setdefault("video_paths", {})[ep] = p; self.set_dirty()

    def change_episode(self):
        ep = self.ep_combo.currentData()
        if not ep: return
        path = self.data["episodes"].get(ep)
        if path and os.path.exists(path): 
            self.table_stack.setCurrentIndex(0)
            self.parse_ass(path)
            self.get_episode_lines(ep)  # Загружаем полные данные строк
            self.refresh_main_table()
            self.update_save_ass_button()  # Обновляем статус кнопки сохранения
        else: 
            self.table_stack.setCurrentIndex(1)

    def import_ass(self, paths=None):
        if not paths: paths, _ = QFileDialog.getOpenFileNames(self, "ASS", "", "*.ass")
        if paths:
            for p in paths:
                num = "".join(re.findall(r'\d+', os.path.basename(p))) or "1"
                n, ok = QInputDialog.getText(self, "Ep", f"Ep for {os.path.basename(p)}:", text=num)
                if ok and n: self.data["episodes"][n] = p; self.parse_ass(p); self.set_dirty()
            self.update_ep_list()

    def relink_file(self):
        ep = self.ep_combo.currentData()
        p, _ = QFileDialog.getOpenFileName(self, "Файл", "", "*.ass")
        if p: self.data["episodes"][ep] = p; self.change_episode(); self.set_dirty()

    def parse_ass(self, path):
        char_data = {}; gap = self.data["export_config"].get('merge_gap', 5)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                lines_list = []
                for line in f:
                    if line.startswith("Dialogue:"):
                        p = line.split(',', 9); c, t = p[4].strip(), re.sub(r'\{.*?\}', '', p[9]).strip()
                        if t: lines_list.append({'s': ass_time_to_seconds(p[1]), 'e': ass_time_to_seconds(p[2]), 'char': c, 'text': t})
                for l in lines_list:
                    if l['char'] not in char_data: char_data[l['char']] = {"lines":0, "raw":[]}
                    char_data[l['char']]["lines"] += 1; char_data[l['char']]["raw"].append(l)
                res = []
                for c, info in char_data.items():
                    r, w, c_l = 1, 0, info["raw"]
                    if c_l:
                        w += len(c_l[0]['text'].split())
                        for i in range(1, len(c_l)):
                            if (c_l[i]['s'] - c_l[i-1]['e']) >= gap: r += 1
                            w += len(c_l[i]['text'].split())
                    res.append({"name": c, "lines": info["lines"], "rings": r, "words": w})
            self.current_ep_stats = res
        except: self.current_ep_stats = []
        
    def update_map(self, char_name, combo):
        aid = combo.currentData()
        if aid: self.data["global_map"][char_name] = aid
        elif char_name in self.data["global_map"]: del self.data["global_map"][char_name]
        self.refresh_actor_list(); self.set_dirty(True)

    def refresh_main_table(self):
        self.main_table.blockSignals(True)
        self.main_table.setRowCount(0); q = self.search_edit.text().lower(); only_p = self.filter_unassigned.isChecked()
        keys = ["name", "lines", "rings", "words"]
        sl = sorted(self.current_ep_stats, key=lambda x: x[keys[self.sort_col]], reverse=self.sort_desc)
        for s in sl:
            if q and q not in s["name"].lower(): continue
            is_a = s["name"] in self.data["global_map"]
            if only_p and is_a: continue
            r = self.main_table.rowCount(); self.main_table.insertRow(r)
            # Создаём редактируемый элемент для имени персонажа
            name_item = QTableWidgetItem(s["name"])
            name_item.setFlags(name_item.flags() | Qt.ItemIsEditable)
            name_item.setData(Qt.UserRole, s["name"])  # Сохраняем оригинальное имя
            self.main_table.setItem(r, 0, name_item)
            self.main_table.setItem(r, 1, QTableWidgetItem(str(s["lines"]))); self.main_table.setItem(r, 2, QTableWidgetItem(str(s["rings"]))); self.main_table.setItem(r, 3, QTableWidgetItem(str(s["words"])))
            cb = QComboBox(); cb.addItem("-", None)
            for aid, info in self.data["actors"].items(): cb.addItem(info["name"], aid)
            if is_a: cb.setCurrentIndex(cb.findData(self.data["global_map"][s["name"]]))
            cb.currentIndexChanged.connect(lambda _, c=s["name"], b=cb: self.update_map(c, b)); self.main_table.setCellWidget(r, 4, cb)
            btn = QPushButton("📺"); btn.setFixedWidth(40); btn.clicked.connect(lambda ch=False, c=s["name"]: self.open_preview(c)); self.main_table.setCellWidget(r, 5, wrap_widget(btn))
        self.main_table.blockSignals(False)

    def refresh_actor_list(self):
        self.actor_table.blockSignals(True); self.actor_table.setRowCount(0); actor_roles = {aid: [] for aid in self.data["actors"]}
        for char, aid in self.data["global_map"].items():
            if aid in actor_roles: actor_roles[aid].append(char)
        for aid, info in self.data["actors"].items():
            row = self.actor_table.rowCount(); self.actor_table.insertRow(row); item = QTableWidgetItem(info["name"]); item.setData(Qt.UserRole, aid); self.actor_table.setItem(row, 0, item); c_it = QTableWidgetItem(); c_it.setBackground(QColor(info["color"])); self.actor_table.setItem(row, 1, c_it); btn = QPushButton(f"Роли ({len(actor_roles[aid])})"); btn.clicked.connect(lambda _, a=aid, n=info["name"], r=actor_roles[aid]: self.edit_roles(a, n, r)); self.actor_table.setCellWidget(row, 2, wrap_widget(btn))
        self.actor_table.blockSignals(False)

    def rename_episode(self):
        old = self.ep_combo.currentData(); n, ok = QInputDialog.getText(self, "Rename", "New name:", text=str(old))
        if ok and n: self.data["episodes"][n] = self.data["episodes"].pop(old); self.update_ep_list(n); self.set_dirty()

    def update_ep_list(self, sel=None):
        self.ep_combo.blockSignals(True); self.ep_combo.clear()
        for ep in sorted(self.data["episodes"].keys(), key=lambda x: int(x) if x.isdigit() else 0): self.ep_combo.addItem(f"Серия {ep}", ep)
        if sel: self.ep_combo.setCurrentIndex(self.ep_combo.findData(sel))
        elif self.ep_combo.count() > 0: self.ep_combo.setCurrentIndex(0)
        self.ep_combo.blockSignals(False); self.change_episode()

    def run_unified_export(self):
        do_html, do_xls = self.chk_exp_html.isChecked(), self.chk_exp_xls.isChecked()
        if not (do_html or do_xls): return
        is_all = self.radio_all.isChecked()
        episodes = self.data["episodes"] if is_all else {self.ep_combo.currentData(): self.data["episodes"].get(self.ep_combo.currentData())}
        if not episodes or None in episodes.values(): return
        if is_all or (do_html and do_xls):
            dest = QFileDialog.getExistingDirectory(self, "Выберите папку")
            if dest: self._execute_batch_export(episodes, do_html, do_xls, dest)
        else:
            ep = list(episodes.keys())[0]
            if do_html: self.export_to_html(ep)
            else: self.export_to_excel(ep)

    def _execute_batch_export(self, episodes, do_html, do_xls, folder):
        cfg = self.data["export_config"]
        highlight_ids = cfg.get('highlight_ids_export')
        for ep, path in episodes.items():
            lines = self.get_episode_lines(ep)
            if not lines: continue
            proc = self.process_merge_logic(lines, cfg)
            if do_html:
                with open(os.path.join(folder, f"{self.data['project_name']} - Ep{ep}.html"), 'w', encoding='utf-8') as f: f.write(self.generate_html_body(ep, proc, cfg, highlight_ids))
            if do_xls and EXCEL_AVAILABLE: self._create_excel_book(ep, proc, cfg).save(os.path.join(folder, f"{self.data['project_name']} - Ep{ep}.xlsx"))
        os.system(f'open "{folder}"') if sys.platform == 'darwin' else os.startfile(folder)

    def _create_excel_book(self, ep, proc, cfg=None):
        if cfg is None:
            cfg = self.data["export_config"]
        
        wb = openpyxl.Workbook(); ws = wb.active; ws.append(["№", "Таймкод", "Персонаж", "Актер", "Текст"])
        use_color = cfg.get('use_color', True)
        highlight_ids = cfg.get('highlight_ids_export')
        all_actor_ids = set(self.data["actors"].keys())
        is_full_filter = highlight_ids is not None and set(highlight_ids) == all_actor_ids
        effective_filter = None if (highlight_ids is None or is_full_filter) else set(highlight_ids)
        
        for i, l in enumerate(proc, 2):
            aid = self.data["global_map"].get(l['char']); act = self.data["actors"].get(aid, {"name": "-", "color": "#FFFFFF"})
            ws.append([i-1, l['s_raw'], l['char'], act['name'], l['text']])
            
            # Применяем цвета только если включено и актёр в фильтре
            if use_color:
                is_h = (effective_filter is None) or (aid in effective_filter)
                if is_h:
                    color = act['color'].replace("#","")
                else:
                    color = "FFFFFF"
            else:
                color = "FFFFFF"
            
            f = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for c in range(1, 6): ws.cell(row=i, column=c).fill = f; ws.cell(row=i, column=c).alignment = Alignment(vertical='top', wrap_text=(c==5))
        return wb

    def export_to_excel(self, ep):
        if not EXCEL_AVAILABLE: return
        wb = self._create_excel_book(ep, self.process_merge_logic(self.get_episode_lines(ep), self.data["export_config"]), self.data["export_config"])
        p, _ = QFileDialog.getSaveFileName(self, "Save Excel", f"Script_{ep}.xlsx", "*.xlsx")
        if p: wb.save(p); os.startfile(p) if sys.platform == 'win32' else os.system(f'open "{p}"')

    def export_to_html(self, ep):
        h = self.generate_html_body(ep, self.process_merge_logic(self.get_episode_lines(ep), self.data["export_config"]), self.data["export_config"], self.data["export_config"].get('highlight_ids_export'))
        p, _ = QFileDialog.getSaveFileName(self, "Save HTML", f"Script_{ep}.html", "*.html")
        if p: 
            with open(p, 'w', encoding='utf-8') as f: f.write(h)
            os.system(f'open "{p}"') if sys.platform == 'darwin' else os.startfile(p)
    
    def save_episode_to_ass(self, ep_num, target_path=None):
        # Убедимся что loaded_episodes загружена
        mem_lines = self.get_episode_lines(ep_num)
        if not mem_lines: 
            QMessageBox.warning(self, "Ошибка", "Нет данных."); 
            return False
        
        source_path = self.data["episodes"].get(ep_num)
        if not source_path or not os.path.exists(source_path): 
            QMessageBox.warning(self, "Ошибка", "Файл не найден."); 
            return False
        
        save_path = target_path if target_path else source_path
        new_file_content = []
        
        try:
            with open(source_path, 'r', encoding='utf-8') as f:
                dia_idx = 0
                for line in f:
                    if line.startswith("Dialogue:"):
                        if dia_idx < len(mem_lines):
                            current_data = mem_lines[dia_idx]
                            parts = line.strip().split(',', 9)
                            if len(parts) > 9:
                                # Обновляем и текст (parts[9]) и имя персонажа (parts[4])
                                parts[4] = current_data['char']  # Новое имя персонажа
                                new_line = f"{','.join(parts[:9])},{current_data['text']}\n"
                                new_file_content.append(new_line)
                            else: 
                                new_file_content.append(line)
                        else: 
                            new_file_content.append(line)
                        dia_idx += 1
                    else: 
                        new_file_content.append(line)
            
            with open(save_path, 'w', encoding='utf-8') as f: 
                f.writelines(new_file_content)
            
            return True
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось записать файл:\n{e}"); 
            return False

    def generate_html_body(self, ep, proc, cfg, highlight_ids=None, override_layout=None):
        layout_mode = override_layout if override_layout else cfg.get('layout_type', "Таблица")
        
        js = """
        <script src="qrc:///qtwebchannel/qwebchannel.js"></script>
        <script>
            var backend;
            new QWebChannel(qt.webChannelTransport, function (channel) {
                backend = channel.objects.backend;
                window.updateScrollStatus();
            });

            window.updateScrollStatus = function() {
                var blocks = Array.from(document.querySelectorAll('.highlighted-block'));
                if (blocks.length === 0 || !backend) return;
                var midLine = window.innerHeight / 2;
                var closestIndex = 0;
                var minDistance = Infinity;
                blocks.forEach((block, index) => {
                    var rect = block.getBoundingClientRect();
                    var distance = Math.abs((rect.top + rect.height/2) - midLine);
                    if (distance < minDistance) {
                        minDistance = distance;
                        closestIndex = index;
                    }
                });
                backend.sync_scroll_index(closestIndex, blocks.length);
            };

            var scrollTimeout;
            window.onscroll = function() {
                clearTimeout(scrollTimeout);
                scrollTimeout = setTimeout(window.updateScrollStatus, 50);
            };
            
            window.jumpToNextHighlighted = function(direction) {
                var blocks = Array.from(document.querySelectorAll('.highlighted-block'));
                if (blocks.length === 0) return;
                
                var targetIndex = -1;
                var threshold = 160; 

                if (direction === 'next') {
                    targetIndex = blocks.findIndex(b => b.getBoundingClientRect().top > threshold);
                    if (targetIndex === -1) targetIndex = 0;
                } else {
                    targetIndex = blocks.findLastIndex(b => b.getBoundingClientRect().top < 50);
                    if (targetIndex === -1) targetIndex = blocks.length - 1;
                }
                
                var target = blocks[targetIndex];
                blocks.forEach(b => b.classList.remove('active-replica'));
                target.classList.add('active-replica');
                target.scrollIntoView({ behavior: 'smooth', block: 'center' });
            };

            function onBlur(el) {
                if(backend) {
                    var cleanText = el.innerText.replace(/(\\r\\n|\\n|\\r)/gm, "");
                    backend.update_text(el.id, cleanText);
                }
            }
            function onKeyPress(e, el) { if (e.keyCode === 13) { e.preventDefault(); el.blur(); } }
        </script>
        <style>
            .edit-span { border-bottom: 1px dashed #ccc; padding: 1px 2px; }
            .edit-span:focus { background-color: #fff; outline: 2px solid #5B9BD5; border-bottom: none; }
            .sep { color: #888; font-weight: bold; }
            .highlighted-block { transition: outline 0.3s, box-shadow 0.3s; }
            .active-replica {
                outline: 6px solid #FFD700 !important;
                outline-offset: -6px;
                box-shadow: 0 0 25px rgba(255, 215, 0, 0.8) !important;
                z-index: 99;
            }
        </style>
        """

        h = f"<html><head><meta charset='utf-8'>{js}<style>"
        h += "body { font-family: 'Segoe UI', sans-serif; padding: 50px 10%; background: #fdfdfd; }"
        h += "table { width: 100%; border-collapse: collapse; table-layout: fixed; background: white; }"
        h += "td, th { border: 1px solid #ddd; padding: 12px; vertical-align: top; overflow-wrap: break-word; }"
        h += f".t {{ width: 90px; font-family: monospace; font-size: {cfg.get('f_time', 12)}px; color: #666; }}"
        h += f".c {{ width: 160px; font-weight: bold; font-size: {cfg.get('f_char', 14)}px; }}"
        h += f".a {{ width: 160px; font-style: italic; font-size: {cfg.get('f_actor', 14)}px; }}"
        h += f".txt {{ font-size: {cfg.get('f_text', 16)}px; line-height: 1.5; }}"
        h += ".line-container { margin-bottom: 30px; padding: 20px; border-left: 8px solid #eee; background: white; }"
        h += "</style></head><body>"
        h += f"<h1>{self.data['project_name']} - Серия {ep}</h1>"

        all_actor_ids = set(self.data["actors"].keys())
        is_full_filter = highlight_ids is not None and set(highlight_ids) == all_actor_ids
        effective_filter = None if (highlight_ids is None or is_full_filter) else set(highlight_ids)
        use_color = cfg.get('use_color', True)

        for l in proc:
            aid = self.data["global_map"].get(l['char'])
            act = self.data["actors"].get(aid, {"name": "-", "color": "#ffffff"})
            
            is_h = (effective_filter is None) or (aid in effective_filter)
            h_class = "highlighted-block" if is_h else ""
            
            # Применяем цвета только если use_color включен
            if use_color and is_h:
                bg_color = hex_to_rgba_string(act['color'], 0.22)
                border_col = act['color']
            else:
                bg_color = "#ffffff"
                border_col = "#eee"

            text_html = ""
            if 'parts' in l:
                for part in l['parts']:
                    if part['sep']: text_html += f"<span class='sep'>{part['sep']}</span>"
                    text_html += f"<span id='{part['id']}' class='edit-span' contenteditable='true' onblur='onBlur(this)' onkeypress='onKeyPress(event, this)'>{part['text']}</span>"
            else: text_html = l['text']

            if layout_mode == "Таблица":
                if proc.index(l) == 0:
                    h += "<table><thead><tr><th>Время</th><th>Персонаж</th><th>Актер</th><th>Текст</th></tr></thead><tbody>"
                h += f"<tr style='background-color:{bg_color}' class='{h_class}'>"
                h += f"<td class='t'>{l['s_raw']}</td><td class='c'>{l['char']}</td>"
                h += f"<td class='a'>{act['name']}</td><td class='txt'>{text_html}</td></tr>"
                if proc.index(l) == len(proc) - 1: h += "</tbody></table>"
            else:
                h += f"<div class='line-container {h_class}' style='background-color:{bg_color}; border-left-color:{border_col}'>"
                h += f"<div class='meta'><span class='c'><b>{l['char']}</b></span> <span class='t'>[{l['s_raw']}]</span> <span class='a'><i>({act['name']})</i></span></div>"
                h += f"<div class='txt'>{text_html}</div></div>"

        return h + "</body></html>"

    def process_merge_logic(self, lines, cfg):
        ps, pl, gap = cfg.get('p_short', 0.5), cfg.get('p_long', 2.0), cfg.get('merge_gap', 5)
        res = []; curr = None
        if lines:
            curr = lines[0].copy(); curr['parts'] = [{'id': lines[0]['id'], 'text': lines[0]['text'], 'sep': ''}]
            for i in range(1, len(lines)):
                nxt = lines[i]; diff = nxt['s'] - curr['e']
                if cfg.get('merge', True) and nxt['char'] == curr['char'] and diff < gap:
                    sep = " // " if diff >= pl else " / " if diff >= ps else " "
                    curr['parts'].append({'id': nxt['id'], 'text': nxt['text'], 'sep': sep})
                    curr['text'] += sep + nxt['text']; curr['e'] = nxt['e']
                else: res.append(curr); curr = nxt.copy(); curr['parts'] = [{'id': nxt['id'], 'text': nxt['text'], 'sep': ''}]
            res.append(curr)
        return res

    def open_preview(self, char):
        ep = self.ep_combo.currentData(); lines = self.get_episode_lines(ep)
        if char: lines = [l for l in lines if l['char'] == char]
        vp = self.data.get("video_paths", {}).get(ep)
        if not vp or not os.path.exists(vp): self.set_episode_video(); vp = self.data.get("video_paths", {}).get(ep)
        if vp: VideoPreviewWindow(vp, lines, ep, self).exec()

    def get_episode_lines(self, ep):
        if "loaded_episodes" not in self.data: self.data["loaded_episodes"] = {}
        if ep in self.data["loaded_episodes"]: return self.data["loaded_episodes"][ep]
        path = self.data["episodes"].get(ep); lines = []
        if path and os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    idx = 0
                    for line in f:
                        if line.startswith("Dialogue:"):
                            p = line.split(',', 9)
                            lines.append({
                                'id': idx, 's': ass_time_to_seconds(p[1]), 'e': ass_time_to_seconds(p[2]),
                                'char': p[4].strip(), 'text': re.sub(r'\{.*?\}', '', p[9]).strip(), 's_raw': p[1]
                            })
                            idx += 1
                self.data["loaded_episodes"][ep] = lines
            except Exception as e: print(f"Read error: {e}")
        return lines

    def show_project_summary(self): SummaryDialog(self.data, None, self).exec()
    def show_episode_summary(self):
        ep = self.ep_combo.currentData()
        if ep: SummaryDialog(self.data, ep, self).exec()
    def edit_roles(self, aid, n, r):
        d = ActorRolesDialog(n, r, self)
        if d.exec():
            new = d.get_roles(); [self.data["global_map"].pop(k, None) for k in [k for k,v in self.data["global_map"].items() if v==aid]]
            for r_name in new: self.data["global_map"][r_name] = aid
            self.refresh_actor_list(); self.refresh_main_table(); self.set_dirty()

    def open_export_settings(self):
        d = ExportSettingsDialog(self.data["export_config"], self)
        if d.exec(): self.data["export_config"] = d.get_settings(); self.change_episode(); self.set_dirty()

    def dragEnterEvent(self, e): e.accept() if e.mimeData().hasUrls() else e.ignore()
    def dropEvent(self, e):
        files = [u.toLocalFile() for u in e.mimeData().urls()]
        ass = [f for f in files if f.endswith('.ass')]; json_f = [f for f in files if f.endswith('.json')]
        if json_f and self.maybe_save(): self._load_from_path(json_f[0])
        elif ass: self.import_ass(ass)

if __name__ == "__main__":
    app = QApplication(sys.argv); w = DubbingApp(); w.show(); sys.exit(app.exec())
